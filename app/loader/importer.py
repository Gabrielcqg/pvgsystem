from __future__ import annotations

import hashlib
import json
import os
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import psycopg
from openpyxl import load_workbook
from psycopg.rows import dict_row

from app.loader.normalization import (
    CATEGORIA_MAP,
    LANCAMENTO_TIPO_MAP,
    PARCELA_TIPO_MAP,
    STATUS_MAP,
    TIPO_HONORARIO_MAP,
    chave_texto,
    map_enum,
    parse_bool,
    parse_date,
    parse_money,
    parse_month,
    parse_percent,
    resolve_headers,
)


MONTH_SHEETS = {
    "JAN": 1,
    "FEV": 2,
    "MAR": 3,
    "ABR": 4,
    "MAI": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SET": 9,
    "OUT": 10,
    "NOV": 11,
    "DEZ": 12,
}


@dataclass
class LoadIssue:
    sheet: str
    row: int | None
    column: str | None
    message: str
    severity: str = "aviso"
    original: str | None = None
    final: str | None = None


@dataclass
class LoadReport:
    execucao_id: UUID
    validate_only: bool
    bindings: dict[str, dict[str, str]] = field(default_factory=dict)
    counts: dict[str, int] = field(default_factory=lambda: {
        "parceiros_criados": 0,
        "contratos": 0,
        "contratos_stub": 0,
        "parcelas": 0,
        "lancamentos": 0,
        "custos_fixos": 0,
        "parametros": 0,
        "linhas_ignoradas_idempotencia": 0,
        "import_log": 0,
    })
    warnings: list[LoadIssue] = field(default_factory=list)
    errors: list[LoadIssue] = field(default_factory=list)
    cash_deltas: dict[str, str] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return not self.errors

    def warn(self, issue: LoadIssue) -> None:
        self.warnings.append(issue)

    def error(self, issue: LoadIssue) -> None:
        issue.severity = "erro"
        self.errors.append(issue)

    def as_dict(self) -> dict[str, Any]:
        return {
            "execucao_id": str(self.execucao_id),
            "validate_only": self.validate_only,
            "ok": self.ok,
            "bindings": self.bindings,
            "counts": self.counts,
            "warnings": [issue.__dict__ for issue in self.warnings],
            "errors": [issue.__dict__ for issue in self.errors],
            "cash_deltas": self.cash_deltas,
        }


def _json_default(value: Any) -> str:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _row_hash(data: dict[str, Any]) -> str:
    payload = json.dumps(data, sort_keys=True, default=_json_default, ensure_ascii=True)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _source_key(sheet: str, kind: str, row: int) -> str:
    return hashlib.sha256(f"{sheet}|{kind}|{row}".encode("utf-8")).hexdigest()


def _cell(row: tuple[Any, ...], columns: dict[str, int], key: str) -> Any:
    col = columns.get(key)
    if not col or col > len(row):
        return None
    return row[col - 1]


def _find_header_row(ws, profile: str, required: set[str], *, start: int = 1, end: int = 25) -> tuple[int, dict[str, int]]:
    last = min(end, ws.max_row)
    best_error: str | None = None
    best_score = -1
    for row_no in range(start, last + 1):
        values = [cell.value for cell in ws[row_no]]
        score = sum(1 for value in values if value not in (None, ""))
        try:
            columns = resolve_headers(values, profile, required)
        except ValueError as exc:
            if score > best_score:
                best_score = score
                best_error = str(exc)
            continue
        return row_no, columns
    if best_error:
        raise ValueError(best_error)
    raise ValueError(f"Nao encontrei cabecalho em {ws.title}; colunas obrigatorias: {sorted(required)}")


def _binding_names(ws, header_row: int, columns: dict[str, int]) -> dict[str, str]:
    return {target: str(ws.cell(header_row, col).value) for target, col in columns.items()}


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(value is None or value == "" for value in row)


def _insert_log(
    cur,
    report: LoadReport,
    *,
    arquivo: str,
    aba: str,
    linha: int | None,
    coluna: str | None,
    acao: str,
    severidade: str = "info",
    valor_origem: Any = None,
    valor_final: Any = None,
    entidade: str | None = None,
    entidade_id: str | UUID | None = None,
    chave_linha: str | None = None,
) -> None:
    cur.execute(
        """
        INSERT INTO import_log (
          execucao_id, arquivo, aba, linha, coluna, acao, valor_origem, valor_final,
          entidade, entidade_id, severidade, chave_linha
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT DO NOTHING
        """,
        (
            str(report.execucao_id),
            arquivo,
            aba,
            linha,
            coluna,
            acao,
            None if valor_origem is None else str(valor_origem),
            None if valor_final is None else str(valor_final),
            entidade,
            str(entidade_id) if entidade_id else None,
            severidade,
            chave_linha,
        ),
    )
    if cur.rowcount:
        report.counts["import_log"] += 1


def _latest_import(cur, *, arquivo: str, aba: str, chave_linha: str) -> dict[str, Any] | None:
    cur.execute(
        """
        SELECT entidade, entidade_id, valor_final
        FROM import_log
        WHERE arquivo = %s
          AND aba = %s
          AND chave_linha = %s
          AND acao IN ('linha_importada', 'linha_atualizada')
        ORDER BY id DESC
        LIMIT 1
        """,
        (arquivo, aba, chave_linha),
    )
    row = cur.fetchone()
    return dict(row) if row else None


def _upsert_with_source(
    cur,
    report: LoadReport,
    *,
    arquivo: str,
    aba: str,
    linha: int,
    chave_linha: str,
    entidade: str,
    data: dict[str, Any],
    insert_sql: str,
    insert_params: tuple[Any, ...],
    update_sql: str | None = None,
    update_params: tuple[Any, ...] = (),
) -> str:
    current_hash = _row_hash(data)
    existing = _latest_import(cur, arquivo=arquivo, aba=aba, chave_linha=chave_linha)
    if existing and existing.get("valor_final") == current_hash:
        report.counts["linhas_ignoradas_idempotencia"] += 1
        return str(existing["entidade_id"])
    if existing and update_sql and existing.get("entidade_id"):
        cur.execute(update_sql, (*update_params, existing["entidade_id"]))
        row_id = str(cur.fetchone()["id"])
        action = "linha_atualizada"
    else:
        cur.execute(insert_sql, insert_params)
        row_id = str(cur.fetchone()["id"])
        action = "linha_importada"
    _insert_log(
        cur,
        report,
        arquivo=arquivo,
        aba=aba,
        linha=linha,
        coluna=None,
        acao=action,
        valor_final=current_hash,
        entidade=entidade,
        entidade_id=row_id,
        chave_linha=chave_linha,
    )
    return row_id


def _ensure_partner(cur, report: LoadReport, *, arquivo: str, aba: str, linha: int, nome: str, chave_linha: str) -> str:
    cur.execute("SELECT id FROM parceiros WHERE nome = %s", (nome,))
    row = cur.fetchone()
    if row:
        return str(row["id"])
    cur.execute("INSERT INTO parceiros (nome, revisar) VALUES (%s, true) RETURNING id", (nome,))
    partner_id = str(cur.fetchone()["id"])
    report.counts["parceiros_criados"] += 1
    _insert_log(
        cur,
        report,
        arquivo=arquivo,
        aba=aba,
        linha=linha,
        coluna="parceiro",
        acao="auto_criar_parceiro",
        severidade="aviso",
        valor_origem=nome,
        valor_final=nome,
        entidade="parceiros",
        entidade_id=partner_id,
        chave_linha=chave_linha,
    )
    return partner_id


def _log_mapping(
    cur,
    report: LoadReport,
    *,
    arquivo: str,
    aba: str,
    linha: int,
    coluna: str,
    original: Any,
    final: str,
    chave_linha: str,
) -> None:
    if original is None or chave_texto(original) == chave_texto(final):
        return
    _insert_log(
        cur,
        report,
        arquivo=arquivo,
        aba=aba,
        linha=linha,
        coluna=coluna,
        acao="mapear_vocabulario",
        valor_origem=original,
        valor_final=final,
        chave_linha=chave_linha,
    )


def _parse_config_and_costs(path: Path, report: LoadReport) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    if "CONFIG" not in wb.sheetnames:
        report.error(LoadIssue("CONFIG", None, None, "Aba CONFIG ausente"))
        return {}, []
    ws = wb["CONFIG"]
    params: dict[str, Any] = {"ano": 2026}
    aliases = {
        alias: target
        for target, values in {
            "caixa_inicial_ano": ("caixa inicial",),
            "meta_caixa_ano": ("meta de caixa", "meta caixa"),
            "meta_recorrencia_mensal": ("meta de recorrencia", "meta recorrencia mensal"),
            "recorrencia_atual": ("recorrencia atual",),
        }.items()
        for alias in values
    }
    for row_no in range(1, min(ws.max_row, 12) + 1):
        label = chave_texto(ws.cell(row_no, 1).value)
        if label in aliases:
            params[aliases[label]] = parse_money(ws.cell(row_no, 2).value)
    for key in ("caixa_inicial_ano", "meta_caixa_ano"):
        if key not in params:
            report.error(LoadIssue("CONFIG", None, key, f"Parametro obrigatorio ausente: {key}"))
    params.setdefault("meta_recorrencia_mensal", Decimal("0.00"))
    params.setdefault("recorrencia_atual", Decimal("0.00"))

    try:
        header_row, columns = _find_header_row(ws, "fixed_costs", {"descricao", "valor_mensal"}, start=5)
        report.bindings["CONFIG.fixed_costs"] = _binding_names(ws, header_row, columns)
    except ValueError as exc:
        report.error(LoadIssue("CONFIG", None, None, str(exc)))
        return params, []

    costs: list[dict[str, Any]] = []
    seen_descriptions: set[str] = set()
    for row_no in range(header_row + 1, ws.max_row + 1):
        row = tuple(cell.value for cell in ws[row_no])
        if _is_empty_row(row):
            continue
        descricao = str(_cell(row, columns, "descricao")).strip()
        key = chave_texto(descricao)
        if key in seen_descriptions:
            issue = LoadIssue("CONFIG", row_no, "descricao", f"Custo fixo possivelmente duplicado: {descricao}", "aviso", descricao, descricao)
            report.warn(issue)
        seen_descriptions.add(key)
        costs.append(
            {
                "row": row_no,
                "data": {
                    "descricao": descricao,
                    "valor_mensal": parse_money(_cell(row, columns, "valor_mensal"), positive=True),
                    "recorrente": parse_bool(_cell(row, columns, "recorrente"), default=True),
                    "dia_vencimento": _cell(row, columns, "dia_vencimento"),
                    "mes_inicio": parse_month(_cell(row, columns, "mes_inicio") or "2026-01", default_year=2026),
                    "mes_fim": None,
                },
                "warning": key in seen_descriptions,
            }
        )
    return params, costs


def _parse_cash(path: Path, report: LoadReport, *, year: int) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    entries: list[dict[str, Any]] = []
    running = Decimal("0.00")
    params, _ = _parse_config_and_costs(path, LoadReport(report.execucao_id, True))
    running += params.get("caixa_inicial_ano", Decimal("0.00"))
    for sheet_name, month in MONTH_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            header_row, columns = _find_header_row(ws, "cash", {"data", "descricao", "tipo", "valor", "categoria"})
            report.bindings[f"{sheet_name}.cash"] = _binding_names(ws, header_row, columns)
        except ValueError as exc:
            report.error(LoadIssue(sheet_name, None, None, str(exc)))
            continue

        typed_balance: Decimal | None = None
        month_entries: list[dict[str, Any]] = []
        for row_no in range(header_row + 1, ws.max_row + 1):
            row = tuple(cell.value for cell in ws[row_no])
            if _is_empty_row(row):
                continue
            try:
                tipo_original = _cell(row, columns, "tipo")
                categoria_original = _cell(row, columns, "categoria")
                tipo = map_enum(tipo_original, LANCAMENTO_TIPO_MAP, field="tipo")
                categoria = map_enum(categoria_original, CATEGORIA_MAP, field="categoria")
                valor = parse_money(_cell(row, columns, "valor"), positive=True)
                data = parse_date(_cell(row, columns, "data"))
                if data is None:
                    raise ValueError("Data obrigatoria ausente")
                saldo_cell = _cell(row, columns, "saldo_digitado")
                if saldo_cell not in (None, ""):
                    typed_balance = parse_money(saldo_cell)
                pago = parse_bool(_cell(row, columns, "pago"), default=True)
                item = {
                    "row": row_no,
                    "sheet": sheet_name,
                    "data": {
                        "data": data,
                        "descricao": str(_cell(row, columns, "descricao")).strip(),
                        "tipo": tipo,
                        "valor": valor,
                        "categoria": categoria,
                        "forma_pagamento": _cell(row, columns, "forma_pagamento"),
                        "pago": pago,
                        "observacoes": _cell(row, columns, "observacoes"),
                        "origem": "manual",
                        "origem_id": None,
                    },
                    "mappings": {
                        "tipo": (tipo_original, tipo),
                        "categoria": (categoria_original, categoria),
                    },
                }
            except ValueError as exc:
                report.error(LoadIssue(sheet_name, row_no, None, str(exc)))
                continue
            month_entries.append(item)
            if pago:
                running += valor if tipo == "entrada" else -valor
        if typed_balance is not None and typed_balance != running:
            delta = running - typed_balance
            report.cash_deltas[sheet_name] = str(delta.quantize(Decimal("0.01")))
            report.warn(LoadIssue(sheet_name, None, "saldo_digitado", f"Saldo digitado diverge do recomputado em {delta}", "aviso", str(typed_balance), str(running)))
        entries.extend(month_entries)
    return entries


def _parse_contracts(path: Path, report: LoadReport, *, year: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    contracts: list[dict[str, Any]] = []
    parcelas: list[dict[str, Any]] = []
    defaults = {"CONTRATOS": "ativo", "ENCERRADOS": "encerrado", "PENDENTES": "proposta"}
    for sheet_name, default_status in defaults.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            header_row, columns = _find_header_row(ws, "contracts", {"cliente", "parceiro", "tipo_honorario"})
            report.bindings[f"{sheet_name}.contracts"] = _binding_names(ws, header_row, columns)
        except ValueError as exc:
            report.error(LoadIssue(sheet_name, None, None, str(exc)))
            continue
        for row_no in range(header_row + 1, ws.max_row + 1):
            row = tuple(cell.value for cell in ws[row_no])
            if _is_empty_row(row):
                continue
            try:
                tipo_original = _cell(row, columns, "tipo_honorario")
                status_original = _cell(row, columns, "status")
                tipo = map_enum(tipo_original, TIPO_HONORARIO_MAP, field="tipo_honorario")
                status = default_status if status_original in (None, "") else map_enum(status_original, STATUS_MAP, field="status")
                data = {
                    "cliente": str(_cell(row, columns, "cliente")).strip(),
                    "parceiro": str(_cell(row, columns, "parceiro")).strip(),
                    "numero_processo": _cell(row, columns, "numero_processo"),
                    "status": status,
                    "tipo_honorario": tipo,
                    "percentual_exito": parse_percent(_cell(row, columns, "percentual_exito")),
                    "percentual_sucumbencia": parse_percent(_cell(row, columns, "percentual_sucumbencia")),
                    "percentual_quota": parse_percent(_cell(row, columns, "percentual_quota")),
                    "honorario_fixo_total": parse_money(_cell(row, columns, "honorario_fixo_total")),
                    "valor_causa": parse_money(_cell(row, columns, "valor_causa")),
                    "apelido_split": _cell(row, columns, "apelido_split"),
                    "observacoes": _cell(row, columns, "observacoes"),
                    "data_proposta": parse_date(_cell(row, columns, "data_proposta")),
                    "data_fechamento": parse_date(_cell(row, columns, "data_fechamento")),
                    "revisar": False,
                }
                contracts.append(
                    {
                        "row": row_no,
                        "sheet": sheet_name,
                        "data": data,
                        "mappings": {
                            "tipo_honorario": (tipo_original, tipo),
                            "status": (status_original, status),
                        },
                    }
                )
            except ValueError as exc:
                report.error(LoadIssue(sheet_name, row_no, None, str(exc)))

    if "MENSAIS" in wb.sheetnames:
        ws = wb["MENSAIS"]
        try:
            header_row, columns = _find_header_row(ws, "parcelas", {"cliente", "tipo", "valor", "mes_esperado"})
            report.bindings["MENSAIS.parcelas"] = _binding_names(ws, header_row, columns)
        except ValueError as exc:
            report.error(LoadIssue("MENSAIS", None, None, str(exc)))
        else:
            for row_no in range(header_row + 1, ws.max_row + 1):
                row = tuple(cell.value for cell in ws[row_no])
                if _is_empty_row(row):
                    continue
                try:
                    tipo_original = _cell(row, columns, "tipo")
                    tipo = map_enum(tipo_original, PARCELA_TIPO_MAP, field="parcela_tipo")
                    recebido = parse_bool(_cell(row, columns, "recebido"), default=False)
                    mes_recebimento = parse_month(_cell(row, columns, "mes_recebimento"), default_year=year) if _cell(row, columns, "mes_recebimento") not in (None, "") else None
                    if recebido and mes_recebimento is None:
                        mes_recebimento = parse_month(_cell(row, columns, "mes_esperado"), default_year=year)
                        report.warn(LoadIssue("MENSAIS", row_no, "mes_recebimento", "Parcela recebida sem mes de recebimento; usando mes esperado", "aviso", None, mes_recebimento.isoformat()))
                    parcelas.append(
                        {
                            "row": row_no,
                            "sheet": "MENSAIS",
                            "data": {
                                "cliente": str(_cell(row, columns, "cliente")).strip(),
                                "tipo": tipo,
                                "valor": parse_money(_cell(row, columns, "valor"), positive=True),
                                "mes_esperado": parse_month(_cell(row, columns, "mes_esperado"), default_year=year),
                                "recebido": recebido,
                                "mes_recebimento": mes_recebimento,
                                "observacoes": _cell(row, columns, "observacoes"),
                            },
                            "mappings": {"tipo": (tipo_original, tipo)},
                        }
                    )
                except ValueError as exc:
                    report.error(LoadIssue("MENSAIS", row_no, None, str(exc)))
    return contracts, parcelas


def _write_log_issues(cur, report: LoadReport, *, arquivo: str, issues: list[LoadIssue]) -> None:
    for issue in issues:
        _insert_log(
            cur,
            report,
            arquivo=arquivo,
            aba=issue.sheet,
            linha=issue.row,
            coluna=issue.column,
            acao="validacao",
            severidade=issue.severity,
            valor_origem=issue.original,
            valor_final=issue.final or issue.message,
        )


def _write_load(conn: psycopg.Connection, report: LoadReport, fluxo_path: Path, contratos_path: Path) -> None:
    params, costs = _parse_config_and_costs(fluxo_path, report)
    entries = _parse_cash(fluxo_path, report, year=int(params.get("ano", 2026)))
    contracts, parcelas = _parse_contracts(contratos_path, report, year=int(params.get("ano", 2026)))
    if report.errors:
        raise ValueError("Importacao invalida")

    fluxo_file = fluxo_path.name
    contratos_file = contratos_path.name
    contract_by_cliente: dict[str, str] = {}

    with conn.cursor(row_factory=dict_row) as cur:
        _write_log_issues(cur, report, arquivo=fluxo_file, issues=report.warnings)

        key = _source_key("CONFIG", "parametros", int(params["ano"]))
        params_hash = _row_hash(params)
        current = _latest_import(cur, arquivo=fluxo_file, aba="CONFIG", chave_linha=key)
        if current and current.get("valor_final") == params_hash:
            report.counts["linhas_ignoradas_idempotencia"] += 1
        else:
            cur.execute(
                """
                INSERT INTO parametros (ano, caixa_inicial_ano, meta_caixa_ano, meta_recorrencia_mensal, recorrencia_atual)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (ano) DO UPDATE SET
                  caixa_inicial_ano = EXCLUDED.caixa_inicial_ano,
                  meta_caixa_ano = EXCLUDED.meta_caixa_ano,
                  meta_recorrencia_mensal = EXCLUDED.meta_recorrencia_mensal,
                  recorrencia_atual = EXCLUDED.recorrencia_atual
                """,
                (
                    params["ano"],
                    params["caixa_inicial_ano"],
                    params["meta_caixa_ano"],
                    params["meta_recorrencia_mensal"],
                    params["recorrencia_atual"],
                ),
            )
            report.counts["parametros"] += 1
            _insert_log(cur, report, arquivo=fluxo_file, aba="CONFIG", linha=None, coluna=None, acao="linha_importada", valor_final=params_hash, entidade="parametros", entidade_id=None, chave_linha=key)

        for cost in costs:
            data = cost["data"]
            key = _source_key("CONFIG", "custo_fixo", cost["row"])
            row_id = _upsert_with_source(
                cur,
                report,
                arquivo=fluxo_file,
                aba="CONFIG",
                linha=cost["row"],
                chave_linha=key,
                entidade="custos_fixos",
                data=data,
                insert_sql="""
                    INSERT INTO custos_fixos (descricao, valor_mensal, recorrente, dia_vencimento, mes_inicio, mes_fim)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                """,
                insert_params=(data["descricao"], data["valor_mensal"], data["recorrente"], data["dia_vencimento"], data["mes_inicio"], data["mes_fim"]),
                update_sql="""
                    UPDATE custos_fixos
                    SET descricao = %s, valor_mensal = %s, recorrente = %s, dia_vencimento = %s, mes_inicio = %s, mes_fim = %s
                    WHERE id = %s
                    RETURNING id
                """,
                update_params=(data["descricao"], data["valor_mensal"], data["recorrente"], data["dia_vencimento"], data["mes_inicio"], data["mes_fim"]),
            )
            if row_id:
                report.counts["custos_fixos"] += 1

        _write_log_issues(cur, report, arquivo=contratos_file, issues=report.warnings)
        for contract in contracts:
            data = dict(contract["data"])
            partner_name = data.pop("parceiro")
            key = _source_key(contract["sheet"], "contrato", contract["row"])
            partner_id = _ensure_partner(cur, report, arquivo=contratos_file, aba=contract["sheet"], linha=contract["row"], nome=partner_name, chave_linha=key)
            for column, (original, final) in contract["mappings"].items():
                _log_mapping(cur, report, arquivo=contratos_file, aba=contract["sheet"], linha=contract["row"], coluna=column, original=original, final=final, chave_linha=key)
            db_data = {**data, "parceiro_id": partner_id}
            row_id = _upsert_with_source(
                cur,
                report,
                arquivo=contratos_file,
                aba=contract["sheet"],
                linha=contract["row"],
                chave_linha=key,
                entidade="contratos",
                data=db_data,
                insert_sql="""
                    INSERT INTO contratos (
                      cliente, parceiro_id, numero_processo, status, tipo_honorario,
                      percentual_exito, percentual_sucumbencia, percentual_quota,
                      honorario_fixo_total, valor_causa, apelido_split, observacoes,
                      data_proposta, data_fechamento, revisar
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """,
                insert_params=(
                    db_data["cliente"],
                    db_data["parceiro_id"],
                    db_data["numero_processo"],
                    db_data["status"],
                    db_data["tipo_honorario"],
                    db_data["percentual_exito"],
                    db_data["percentual_sucumbencia"],
                    db_data["percentual_quota"],
                    db_data["honorario_fixo_total"],
                    db_data["valor_causa"],
                    db_data["apelido_split"],
                    db_data["observacoes"],
                    db_data["data_proposta"],
                    db_data["data_fechamento"],
                    db_data["revisar"],
                ),
                update_sql="""
                    UPDATE contratos
                    SET cliente = %s, parceiro_id = %s, numero_processo = %s, status = %s,
                        tipo_honorario = %s, percentual_exito = %s, percentual_sucumbencia = %s,
                        percentual_quota = %s, honorario_fixo_total = %s, valor_causa = %s,
                        apelido_split = %s, observacoes = %s, data_proposta = %s,
                        data_fechamento = %s, revisar = %s
                    WHERE id = %s
                    RETURNING id
                """,
                update_params=(
                    db_data["cliente"],
                    db_data["parceiro_id"],
                    db_data["numero_processo"],
                    db_data["status"],
                    db_data["tipo_honorario"],
                    db_data["percentual_exito"],
                    db_data["percentual_sucumbencia"],
                    db_data["percentual_quota"],
                    db_data["honorario_fixo_total"],
                    db_data["valor_causa"],
                    db_data["apelido_split"],
                    db_data["observacoes"],
                    db_data["data_proposta"],
                    db_data["data_fechamento"],
                    db_data["revisar"],
                ),
            )
            contract_by_cliente[chave_texto(db_data["cliente"])] = row_id
            report.counts["contratos"] += 1

        for parcela in parcelas:
            data = dict(parcela["data"])
            cliente = data.pop("cliente")
            key = _source_key(parcela["sheet"], "parcela", parcela["row"])
            contrato_id = contract_by_cliente.get(chave_texto(cliente))
            if contrato_id is None:
                cur.execute("SELECT id FROM contratos WHERE cliente = %s ORDER BY criado_em LIMIT 1", (cliente,))
                existing = cur.fetchone()
                contrato_id = str(existing["id"]) if existing else None
            if contrato_id is None:
                partner_id = _ensure_partner(cur, report, arquivo=contratos_file, aba=parcela["sheet"], linha=parcela["row"], nome="Pavageau", chave_linha=key)
                cur.execute(
                    """
                    INSERT INTO contratos (cliente, parceiro_id, status, tipo_honorario, revisar, observacoes)
                    VALUES (%s, %s, 'proposta', 'fixo_mensal', true, %s)
                    RETURNING id
                    """,
                    (cliente, partner_id, "Criado automaticamente pelo reconciliador de MENSAIS"),
                )
                created_stub = cur.fetchone()
                if created_stub is None:
                    raise RuntimeError("Falha ao criar contrato stub")
                contrato_id = str(created_stub["id"])
                contract_by_cliente[chave_texto(cliente)] = contrato_id
                report.counts["contratos_stub"] += 1
                _insert_log(
                    cur,
                    report,
                    arquivo=contratos_file,
                    aba=parcela["sheet"],
                    linha=parcela["row"],
                    coluna="cliente",
                    acao="auto_criar_contrato_stub",
                    severidade="aviso",
                    valor_origem=cliente,
                    valor_final=contrato_id,
                    entidade="contratos",
                    entidade_id=contrato_id,
                    chave_linha=key,
                )
            for column, (original, final) in parcela["mappings"].items():
                _log_mapping(cur, report, arquivo=contratos_file, aba=parcela["sheet"], linha=parcela["row"], coluna=column, original=original, final=final, chave_linha=key)
            db_data = {**data, "contrato_id": contrato_id}
            row_id = _upsert_with_source(
                cur,
                report,
                arquivo=contratos_file,
                aba=parcela["sheet"],
                linha=parcela["row"],
                chave_linha=key,
                entidade="parcelas",
                data=db_data,
                insert_sql="""
                    INSERT INTO parcelas (contrato_id, tipo, valor, mes_esperado, recebido, mes_recebimento, observacoes)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """,
                insert_params=(db_data["contrato_id"], db_data["tipo"], db_data["valor"], db_data["mes_esperado"], db_data["recebido"], db_data["mes_recebimento"], db_data["observacoes"]),
                update_sql="""
                    UPDATE parcelas
                    SET contrato_id = %s, tipo = %s, valor = %s, mes_esperado = %s, recebido = %s, mes_recebimento = %s, observacoes = %s
                    WHERE id = %s
                    RETURNING id
                """,
                update_params=(db_data["contrato_id"], db_data["tipo"], db_data["valor"], db_data["mes_esperado"], db_data["recebido"], db_data["mes_recebimento"], db_data["observacoes"]),
            )
            if row_id:
                report.counts["parcelas"] += 1

        for entry in entries:
            data = entry["data"]
            key = _source_key(entry["sheet"], "lancamento", entry["row"])
            for column, (original, final) in entry["mappings"].items():
                _log_mapping(cur, report, arquivo=fluxo_file, aba=entry["sheet"], linha=entry["row"], coluna=column, original=original, final=final, chave_linha=key)
            row_id = _upsert_with_source(
                cur,
                report,
                arquivo=fluxo_file,
                aba=entry["sheet"],
                linha=entry["row"],
                chave_linha=key,
                entidade="lancamentos",
                data=data,
                insert_sql="""
                    INSERT INTO lancamentos (data, descricao, tipo, valor, categoria, forma_pagamento, pago, observacoes, origem, origem_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'manual', NULL)
                    RETURNING id
                """,
                insert_params=(data["data"], data["descricao"], data["tipo"], data["valor"], data["categoria"], data["forma_pagamento"], data["pago"], data["observacoes"]),
                update_sql="""
                    UPDATE lancamentos
                    SET data = %s, descricao = %s, tipo = %s, valor = %s, categoria = %s,
                        forma_pagamento = %s, pago = %s, observacoes = %s
                    WHERE id = %s
                    RETURNING id
                """,
                update_params=(data["data"], data["descricao"], data["tipo"], data["valor"], data["categoria"], data["forma_pagamento"], data["pago"], data["observacoes"]),
            )
            if row_id:
                report.counts["lancamentos"] += 1
        cur.execute("SELECT private.recalcular_mes(%s::smallint, 1::smallint)", (int(params["ano"]),))


def load_workbooks(
    *,
    fluxo_path: str | Path,
    contratos_path: str | Path,
    database_url: str | None = None,
    conn: psycopg.Connection | None = None,
    validate_only: bool = True,
    execucao_id: UUID | None = None,
) -> LoadReport:
    fluxo = Path(fluxo_path)
    contratos = Path(contratos_path)
    report = LoadReport(execucao_id or uuid4(), validate_only)

    params, _costs = _parse_config_and_costs(fluxo, report)
    _parse_cash(fluxo, report, year=int(params.get("ano", 2026)))
    _parse_contracts(contratos, report, year=int(params.get("ano", 2026)))
    if validate_only:
        return report
    if report.errors:
        return report

    owns_conn = conn is None
    if conn is None:
        url = database_url or os.getenv("DATABASE_URL")
        if not url:
            raise ValueError("DATABASE_URL e obrigatorio para --commit")
        conn = psycopg.connect(url)
    try:
        with conn.transaction():
            _write_load(conn, report, fluxo, contratos)
    finally:
        if owns_conn:
            conn.close()
    return report
