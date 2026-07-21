from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook


@dataclass(frozen=True)
class SyntheticFixturePaths:
    fluxo: Path
    contratos: Path
    esperado: Path


def _write_rows(ws, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)


def generate_synthetic_fixtures(target_dir: str | Path) -> SyntheticFixturePaths:
    target = Path(target_dir)
    target.mkdir(parents=True, exist_ok=True)
    fluxo_path = target / "fluxo_caixa_sintetico.xlsx"
    contratos_path = target / "contratos_sintetico.xlsx"
    esperado_path = target / "esperado.json"

    fluxo = Workbook()
    config = fluxo.active
    config.title = "CONFIG"
    _write_rows(
        config,
        [
            ["Caixa inicial", 1000],
            ["Meta caixa", 5000],
            ["Meta recorrência mensal", 2000],
            ["Recorrência atual", 1500],
            [],
            ["Descrição", "Valor mensal", "Recorrente", "Dia", "Mês início"],
            ["Contador", 500, True, 5, "2026-01"],
            ["Contador", 500, True, 5, "2026-01"],
            ["Software", 100, True, 10, "2026-01"],
        ],
    )
    month_rows = {
        "JAN": [
            ["2026-01-05", "Honorário Cliente 01", "Entrada", 2000, "Honorários", "PIX", "sim", "ok", 9999],
            ["2026-01-10", "Infraestrutura", "Saída", 300, "Infraestrutura", "Boleto", "sim", "ok", None],
        ],
        "FEV": [
            ["2026-02-03", "Consultoria", "E", 1500, "Consultoria", "PIX", "sim", "", 8888],
            ["2026-02-14", "Marketing", "S", 200, "Marketing", "Cartão", "sim", "", None],
        ],
        "MAR": [
            ["2026-03-20", "Reembolso Cliente", "Saída", 100, "Restituição Cliente", "TED", "sim", "", 7777],
        ],
    }
    for sheet in ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]:
        ws = fluxo.create_sheet(sheet)
        ws.append(["Data", "Histórico", "E/S", "Valor", "Categoria", "Pagamento", "Status", "Observações", "Saldo digitado"])
        for row in month_rows.get(sheet, []):
            ws.append(row)
    dashboard = fluxo.create_sheet("DASHBOARD")
    dashboard.append(["Ignorado", "Derivado"])
    fluxo.save(fluxo_path)

    contracts = Workbook()
    contracts.active.title = "CONTRATOS"
    header = [
        "Cliente",
        "Parceiro",
        "Processo",
        "Tipo de honorário",
        "% êxito",
        "% sucumbência",
        "% quota",
        "Honorário fixo",
        "Valor da causa",
        "Split",
        "Obs",
        "Data proposta",
        "Data fechamento",
        "Situação",
    ]
    spellings = [
        "fixo único",
        "fixo unico",
        "único",
        "fixo mensal",
        "mensal",
        "fixo parcelado",
        "parcelado",
        "êxito puro",
        "só êxito",
        "sucumbência",
        "fixo + êxito",
        "fixo e exito",
        "fixo+exito",
        "exito + sucumbencia",
        "fixo + exito + sucumbencia",
    ]
    statuses = ["Aguardando êxito", "Aguardando exito", "ativo", "em andamento", None]
    for sheet_name in ["CONTRATOS", "ENCERRADOS", "PENDENTES"]:
        ws = contracts[sheet_name] if sheet_name in contracts.sheetnames else contracts.create_sheet(sheet_name)
        ws.append(header)
    for idx, spelling in enumerate(spellings, start=1):
        contracts["CONTRATOS"].append(
            [
                f"Cliente {idx:02d}",
                "Parceiro Invisível" if idx == 3 else "Pavageau",
                f"0000000-{idx:02d}.2026.8.26.0001",
                spelling,
                "30%",
                "10%",
                "20%",
                1000,
                10000,
                "",
                "",
                "2026-01-01",
                "2026-01-15",
                statuses[(idx - 1) % len(statuses)],
            ]
        )
    contracts["ENCERRADOS"].append(
        ["Cliente Encerrado", "Pavageau", "", "sucumbência", "0%", "10%", "0%", 0, 5000, "", "", "2026-01-01", "2026-02-01", "sem êxito"]
    )
    contracts["PENDENTES"].append(
        ["Cliente Proposta", "Pavageau", "", "fixo mensal", "0%", "0%", "0%", 1200, 0, "", "", "2026-03-01", "", "em negociação"]
    )
    mensais = contracts.create_sheet("MENSAIS")
    mensais.append(["Contrato", "Tipo", "Valor", "Competência", "Recebido", "Recebido em", "Obs"])
    mensais.append(["Cliente 01", "mensal", 300, "2026-01", True, None, "forca fallback mes_recebimento"])
    for idx in range(1, 5):
        mensais.append([f"Órfão Mensal {idx}", "mensal", 100 + idx, "2026-02", False, None, "sem contrato"])
    contracts.create_sheet("DASHBOARD").append(["Ignorado"])
    contracts.save(contratos_path)

    expected = {
        "counts": {
            "contratos_total": 21,
            "contratos_stub": 4,
            "parcelas": 5,
            "lancamentos": 5,
            "custos_fixos": 3,
            "parceiros_auto_criados": 1,
        },
        "tipo_honorario_distribution": {
            "fixo_unico": 3,
            "fixo_mensal": 7,
            "fixo_parcelado": 2,
            "exito_puro": 2,
            "sucumbencia": 2,
            "fixo_exito": 3,
            "exito_sucumbencia": 1,
            "fixo_exito_sucumbencia": 1,
        },
        "cash_chain": {
            "JAN": "2700.00",
            "FEV": "4000.00",
            "MAR": "3900.00",
            "DEZ": "3900.00",
        },
    }
    esperado_path.write_text(json.dumps(expected, indent=2, ensure_ascii=False), encoding="utf-8")
    return SyntheticFixturePaths(fluxo_path, contratos_path, esperado_path)
