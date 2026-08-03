#!/usr/bin/env python3
"""Import Pavageau process spreadsheet into the operational database.

The importer intentionally treats the spreadsheet as process-cadastro data only.
It does not create clients, contracts, radar movements, radar executions, or
tasks because the source columns do not identify those entities safely.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import psycopg
from openpyxl import load_workbook
from psycopg.conninfo import make_conninfo
from psycopg.rows import dict_row
from psycopg.types.json import Jsonb


DEFAULT_FILE = Path("/Users/gabrielcamargo/Downloads/Pavageau_Top3_Tribunais.xlsx")
DEFAULT_SHEET = "Processos dos 3 Tribunais"
PROCESS_NUMBER_RE = re.compile(r"^\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}(/[0-9]+)?$")
SUPPORTED_RADAR_TRIBUNAIS = {"TJSP"}
TRIBUNAIS_PERMITIDOS = {"TJSP", "TJCE", "TJBA"}

HEADER_MAP = {
    "AREA PASTA": "area_pasta",
    "N": "numero_interno",
    "NO": "numero_interno",
    "PROCESSO N CNJ": "numero",
    "PROCESSO NO CNJ": "numero",
    "TRIBUNAL": "tribunal",
    "STATUS": "status_processo",
    "AUTOR": "autor",
    "REU": "reu",
    "VARA JUIZO": "comarca_vara",
    "ASSUNTO": "assunto",
    "ANDAMENTO ATUAL": "andamento_atual",
}

OPERATIONAL_TABLES = [
    "public.auditoria",
    "public.import_log",
    "public.ind_analise_mensal",
    "public.ind_balanco",
    "public.ind_dre_mensal",
    "public.ind_fluxo_mensal",
    "public.ind_gastos_categoria",
    "public.ind_painel",
    "public.tarefa_dependencias",
    "public.tarefa_tag_relacoes",
    "public.tarefa_status_tempos",
    "public.tarefa_checklist_itens",
    "public.tarefa_subtarefas",
    "public.tarefa_comentarios",
    "public.tarefa_colaboradores",
    "public.tarefas",
    "public.radar_automacao_execucoes",
    "public.radar_movimentacao_classificacoes",
    "public.movimentacoes_novas",
    "public.resultados_consulta",
    "public.execucoes_radar",
    "public.parcelas",
    "public.lancamentos",
    "public.contratos",
    "public.custos_fixos",
    "public.parceiros",
    "public.processos",
]

BACKUP_TABLES = [
    "auditoria",
    "import_log",
    "ind_analise_mensal",
    "ind_balanco",
    "ind_dre_mensal",
    "ind_fluxo_mensal",
    "ind_gastos_categoria",
    "ind_painel",
    "tarefa_dependencias",
    "tarefa_tag_relacoes",
    "tarefa_status_tempos",
    "tarefa_checklist_itens",
    "tarefa_subtarefas",
    "tarefa_comentarios",
    "tarefa_colaboradores",
    "tarefas",
    "radar_automacao_execucoes",
    "radar_movimentacao_classificacoes",
    "movimentacoes_novas",
    "resultados_consulta",
    "execucoes_radar",
    "parcelas",
    "lancamentos",
    "contratos",
    "custos_fixos",
    "parceiros",
    "processos",
]


@dataclass
class ImportRow:
    source_row: int
    area_pasta: str | None
    numero_interno: str | None
    numero: str
    tribunal: str
    status_processo: str | None
    autor: str | None
    reu: str | None
    comarca_vara: str | None
    assunto: str | None
    andamento_atual: str | None
    ativo: bool
    monitorar: bool


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip("\"").strip("'"))


def normalize_header(value: Any) -> str:
    text = normalize_text(value) or ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()
    return re.sub(r"\s+", " ", text)


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text or None


def normalize_cnj(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    return text.replace(" ", "")


def normalize_tribunal(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None
    return text.upper().replace(" ", "")


def is_processo_ativo(status: str | None) -> bool:
    if not status:
        return True
    status_norm = unicodedata.normalize("NFKD", status.lower())
    status_norm = "".join(ch for ch in status_norm if not unicodedata.combining(ch))
    return status_norm not in {"extinto", "baixado", "arquivado", "encerrado", "inativo"}


def find_header(ws) -> tuple[int, dict[int, str]]:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        mapped: dict[int, str] = {}
        for col_idx, value in enumerate(row, start=1):
            key = HEADER_MAP.get(normalize_header(value))
            if key:
                mapped[col_idx] = key
        if {"area_pasta", "numero", "tribunal", "status_processo"}.issubset(set(mapped.values())):
            return row_idx, mapped
    raise RuntimeError("Cabecalho da planilha nao encontrado.")


def parse_workbook(path: Path, sheet_name: str) -> tuple[list[ImportRow], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba {sheet_name!r} nao encontrada. Abas: {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row, col_map = find_header(ws)
    valid: list[ImportRow] = []
    issues: list[dict[str, Any]] = []
    seen: set[str] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        raw = {field: row[col_idx - 1] if col_idx - 1 < len(row) else None for col_idx, field in col_map.items()}
        if not any(normalize_text(value) for value in raw.values()):
            continue
        numero = normalize_cnj(raw.get("numero"))
        tribunal = normalize_tribunal(raw.get("tribunal"))
        if not numero:
            issues.append({"linha": row_idx, "tipo": "invalido", "motivo": "PROCESSO (Nº CNJ) vazio"})
            continue
        if not PROCESS_NUMBER_RE.fullmatch(numero):
            issues.append({"linha": row_idx, "tipo": "invalido", "cnj": numero, "motivo": "Numero processual fora do formato esperado"})
            continue
        if not tribunal or tribunal not in TRIBUNAIS_PERMITIDOS:
            issues.append({"linha": row_idx, "tipo": "invalido", "cnj": numero, "tribunal": tribunal, "motivo": "Tribunal nao permitido"})
            continue
        if numero in seen:
            issues.append({"linha": row_idx, "tipo": "duplicado_arquivo", "cnj": numero, "motivo": "CNJ repetido no arquivo"})
            continue
        seen.add(numero)
        status = normalize_text(raw.get("status_processo"))
        ativo = is_processo_ativo(status)
        valid.append(
            ImportRow(
                source_row=row_idx,
                area_pasta=normalize_text(raw.get("area_pasta")),
                numero_interno=normalize_text(raw.get("numero_interno")),
                numero=numero,
                tribunal=tribunal,
                status_processo=status,
                autor=normalize_text(raw.get("autor")),
                reu=normalize_text(raw.get("reu")),
                comarca_vara=normalize_text(raw.get("comarca_vara")),
                assunto=normalize_text(raw.get("assunto")),
                andamento_atual=normalize_text(raw.get("andamento_atual")),
                ativo=ativo,
                monitorar=ativo and tribunal in SUPPORTED_RADAR_TRIBUNAIS,
            )
        )
    return valid, issues


def json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, UUID):
        return str(value)
    return str(value)


def env_dsn() -> str:
    dsn = os.getenv("MIGRATION_DATABASE_URL") or os.getenv("DATABASE_URL")
    if not dsn:
        raise RuntimeError("MIGRATION_DATABASE_URL ou DATABASE_URL precisa estar configurado.")
    return make_conninfo(dsn, connect_timeout="12")


def fetch_table_counts(conn: psycopg.Connection, tables: list[str]) -> dict[str, int]:
    counts: dict[str, int] = {}
    with conn.cursor() as cur:
        for table in tables:
            cur.execute(f"SELECT count(*) FROM public.{table}")
            row = cur.fetchone()
            if row is None:
                raise RuntimeError(f"Contagem sem retorno para {table}")
            counts[table] = int(row[0])
    return counts


def required_row(row: dict[str, Any] | None, context: str) -> dict[str, Any]:
    if row is None:
        raise RuntimeError(f"Consulta sem retorno: {context}")
    return row


def backup_operational_data(conn: psycopg.Connection, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    backup_path = output_dir / "backup_operational_before_import.json"
    data: dict[str, Any] = {"created_at": datetime.now().isoformat(), "tables": {}}
    with conn.cursor(row_factory=dict_row) as cur:
        for table in BACKUP_TABLES:
            cur.execute(f"SELECT * FROM public.{table}")
            data["tables"][table] = [dict(row) for row in cur.fetchall()]
    backup_path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=json_default))
    return backup_path


def cleanup_operational_data(conn: psycopg.Connection) -> None:
    tables_sql = ", ".join(OPERATIONAL_TABLES)
    with conn.cursor() as cur:
        cur.execute(f"TRUNCATE TABLE {tables_sql} RESTART IDENTITY CASCADE")


def insert_processes(conn: psycopg.Connection, rows: list[ImportRow], arquivo: Path, aba: str, batch_id: UUID) -> tuple[int, int]:
    inserted_or_updated = 0
    monitorados = 0
    with conn.cursor(row_factory=dict_row) as cur:
        for row in rows:
            payload = asdict(row)
            cur.execute(
                """
                INSERT INTO public.processos (
                  area_pasta, numero_interno, numero, cliente, tribunal, status_processo,
                  autor, reu, comarca_vara, assunto, andamento_atual,
                  ativo, monitorar, contrato_id, data_ultimo_andamento, chaves_movimentacoes,
                  exige_senha, ultima_consulta_status, ultima_consulta_em, ultima_consulta_inconclusiva
                )
                VALUES (
                  %(area_pasta)s, %(numero_interno)s, %(numero)s, NULL, %(tribunal)s, %(status_processo)s,
                  %(autor)s, %(reu)s, %(comarca_vara)s, %(assunto)s, %(andamento_atual)s,
                  %(ativo)s, %(monitorar)s, NULL, NULL, '{}'::text[],
                  false, NULL, NULL, false
                )
                ON CONFLICT (numero) DO UPDATE SET
                  area_pasta = excluded.area_pasta,
                  numero_interno = excluded.numero_interno,
                  cliente = NULL,
                  tribunal = excluded.tribunal,
                  status_processo = excluded.status_processo,
                  autor = excluded.autor,
                  reu = excluded.reu,
                  comarca_vara = excluded.comarca_vara,
                  assunto = excluded.assunto,
                  andamento_atual = excluded.andamento_atual,
                  ativo = excluded.ativo,
                  monitorar = excluded.monitorar,
                  contrato_id = NULL,
                  data_ultimo_andamento = NULL,
                  chaves_movimentacoes = '{}'::text[],
                  exige_senha = false,
                  ultima_consulta_status = NULL,
                  ultima_consulta_em = NULL,
                  ultima_consulta_inconclusiva = false
                RETURNING id
                """,
                payload,
            )
            processo_id = required_row(cur.fetchone(), "processo importado")["id"]
            inserted_or_updated += 1
            if row.monitorar:
                monitorados += 1
            cur.execute(
                """
                INSERT INTO public.import_log (
                  execucao_id, arquivo, aba, linha, acao, valor_origem, valor_final,
                  entidade, entidade_id, severidade, chave_linha
                )
                VALUES (%s, %s, %s, %s, 'importar_processo', %s, %s, 'processos', %s, 'info', %s)
                """,
                (
                    batch_id,
                    str(arquivo),
                    aba,
                    row.source_row,
                    row.numero,
                    json.dumps(payload, ensure_ascii=False, default=json_default),
                    processo_id,
                    row.numero,
                ),
            )
    return inserted_or_updated, monitorados


def log_issues(conn: psycopg.Connection, issues: list[dict[str, Any]], arquivo: Path, aba: str, batch_id: UUID) -> None:
    if not issues:
        return
    with conn.cursor() as cur:
        for issue in issues:
            cur.execute(
                """
                INSERT INTO public.import_log (
                  execucao_id, arquivo, aba, linha, coluna, acao, valor_origem, valor_final,
                  entidade, severidade, chave_linha
                )
                VALUES (%s, %s, %s, %s, %s, 'ignorar_linha', %s, %s, 'processos', %s, %s)
                """,
                (
                    batch_id,
                    str(arquivo),
                    aba,
                    issue.get("linha"),
                    issue.get("coluna"),
                    json.dumps(issue, ensure_ascii=False),
                    issue.get("motivo"),
                    "warning" if issue.get("tipo", "").startswith("duplicado") else "error",
                    issue.get("cnj"),
                ),
            )


def validate_import(conn: psycopg.Connection) -> dict[str, Any]:
    with conn.cursor(row_factory=dict_row) as cur:
        cur.execute("SELECT count(*) AS total FROM public.processos")
        total = int(required_row(cur.fetchone(), "total processos")["total"])
        cur.execute("SELECT tribunal::text AS tribunal, count(*) AS total FROM public.processos GROUP BY tribunal ORDER BY tribunal")
        por_tribunal = {row["tribunal"]: int(row["total"]) for row in cur.fetchall()}
        cur.execute("SELECT count(*) AS total FROM public.processos WHERE coalesce(cardinality(chaves_movimentacoes), 0) > 0")
        processos_com_base = int(required_row(cur.fetchone(), "processos com base radar")["total"])
        cur.execute("SELECT count(*) AS total FROM public.movimentacoes_novas")
        movimentacoes_novas = int(required_row(cur.fetchone(), "movimentacoes novas")["total"])
        cur.execute("SELECT count(*) AS total FROM public.resultados_consulta")
        resultados_consulta = int(required_row(cur.fetchone(), "resultados consulta")["total"])
        cur.execute("SELECT count(*) AS total FROM public.execucoes_radar")
        execucoes_radar = int(required_row(cur.fetchone(), "execucoes radar")["total"])
        cur.execute("SELECT count(*) AS total FROM public.tarefas")
        tarefas = int(required_row(cur.fetchone(), "tarefas")["total"])
        cur.execute("SELECT count(*) AS total FROM public.processos WHERE tribunal::text <> 'TJSP' AND monitorar IS TRUE")
        unsupported_monitorar = int(required_row(cur.fetchone(), "unsupported monitorar")["total"])
        cur.execute("SELECT count(*) AS total FROM public.processos WHERE tribunal::text = 'TJSP' AND monitorar IS TRUE")
        tjsp_monitorados = int(required_row(cur.fetchone(), "tjsp monitorados")["total"])
        cur.execute("SELECT count(*) AS total FROM public.processos WHERE cliente IS NOT NULL OR contrato_id IS NOT NULL")
        vinculados_cliente_contrato = int(required_row(cur.fetchone(), "vinculos cliente contrato")["total"])
    return {
        "processos_total": total,
        "processos_por_tribunal": por_tribunal,
        "tjsp_monitorados": tjsp_monitorados,
        "unsupported_monitorar": unsupported_monitorar,
        "processos_com_base_radar": processos_com_base,
        "movimentacoes_novas": movimentacoes_novas,
        "resultados_consulta": resultados_consulta,
        "execucoes_radar": execucoes_radar,
        "tarefas": tarefas,
        "vinculados_cliente_contrato": vinculados_cliente_contrato,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa processos Pavageau a partir de planilha Excel.")
    parser.add_argument("--file", type=Path, default=DEFAULT_FILE)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--execute", action="store_true", help="Grava no banco. Sem esta flag, roda dry-run.")
    parser.add_argument("--cleanup", action="store_true", help="Remove dados operacionais antes de importar.")
    parser.add_argument("--backup-dir", type=Path, default=Path("docs/evidence/imports") / datetime.now().strftime("%Y%m%d_%H%M%S"))
    args = parser.parse_args()

    load_dotenv(Path(".env"))
    rows, issues = parse_workbook(args.file, args.sheet)
    found_rows = len(rows) + len(issues)
    by_tribunal = Counter(row.tribunal for row in rows)
    invalid = [issue for issue in issues if issue.get("tipo") == "invalido"]
    duplicates = [issue for issue in issues if issue.get("tipo", "").startswith("duplicado")]

    summary: dict[str, Any] = {
        "arquivo": str(args.file),
        "aba": args.sheet,
        "modo": "execute" if args.execute else "dry-run",
        "linhas_encontradas": found_rows,
        "validas_para_importacao": len(rows),
        "invalidas": len(invalid),
        "duplicadas": len(duplicates),
        "por_tribunal_arquivo": dict(sorted(by_tribunal.items())),
        "monitorar_tjsp_ativos": sum(1 for row in rows if row.monitorar),
        "unsupported_sem_monitoramento": sum(1 for row in rows if row.tribunal not in SUPPORTED_RADAR_TRIBUNAIS),
        "issues": issues,
    }

    if not args.execute:
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=json_default))
        return 0

    batch_id = uuid4()
    with psycopg.connect(env_dsn(), autocommit=False) as conn:
        try:
            before_counts = fetch_table_counts(conn, [table.split(".")[-1] for table in OPERATIONAL_TABLES])
            backup_path = backup_operational_data(conn, args.backup_dir)
            if args.cleanup:
                cleanup_operational_data(conn)
            imported, monitorados = insert_processes(conn, rows, args.file, args.sheet, batch_id)
            log_issues(conn, issues, args.file, args.sheet, batch_id)
            validation = validate_import(conn)
            conn.commit()
        except Exception:
            conn.rollback()
            raise

    summary.update(
        {
            "backup": str(backup_path),
            "cleanup_executado": bool(args.cleanup),
            "batch_id": str(batch_id),
            "contagens_antes": before_counts,
            "importados": imported,
            "monitorados": monitorados,
            "validacao": validation,
        }
    )
    args.backup_dir.mkdir(parents=True, exist_ok=True)
    report_path = args.backup_dir / "import_report.json"
    report_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=json_default))
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=json_default))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
