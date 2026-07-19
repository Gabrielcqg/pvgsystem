from __future__ import annotations

import json
import logging
import random
import re
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from DrissionPage import ChromiumOptions, ChromiumPage
from openpyxl import load_workbook


# ============================================================
# AJUSTE ESTAS CONFIGURAÇÕES
# ============================================================

ARQUIVO_EXCEL = Path(
    "/Users/gabrielcamargo/Downloads/processos.xlsx"
)
NOME_ABA_EXCEL = "Página1"
NOME_COLUNA_PROCESSO = "PROCESSO"
ARQUIVO_JSON = Path("resultado_processos_tjsp.json")

MAX_TENTATIVAS_POR_PROCESSO = 2
RETOMAR_EXECUCAO_ANTERIOR = True

TIMEOUT_FORMULARIO = 120.0
TIMEOUT_RESULTADO = 120.0
WATCHDOG_ACAO = 10.0
INTERVALO_VERIFICACAO = 0.35
INTERVALO_LOG = 3.0
MAX_MOVIMENTACOES = 3

# Pausas de estabilidade. Não burlam nem resolvem captcha.
PAUSA_APOS_ABRIR_ABA = (0.6, 1.2)
PAUSA_ANTES_DE_PREENCHER = (0.4, 0.9)
PAUSA_ANTES_DE_CONSULTAR = (1.8, 2.4)
PAUSA_ENTRE_TENTATIVAS = (1.5, 3.0)
PAUSA_ENTRE_PROCESSOS = (0.8, 1.8)

PASTA_LOGS = Path("logs")
PASTA_DEBUG = Path("debug_tjsp")
ARQUIVO_LOG = PASTA_LOGS / "consulta_processos_tjsp.log"


# ============================================================
# PERFIL DO PORTAL
# Adicione outros perfis quando outro site tiver formulário diferente.
# ============================================================


@dataclass(frozen=True)
class PortalConfig:
    nome: str
    url_consulta: str
    campos_processo: tuple[str, ...]
    botoes_consultar: tuple[str, ...]
    botoes_voltar: tuple[str, ...]
    mensagens_captcha: tuple[str, ...]
    mensagens_senha_necessaria: tuple[str, ...]
    mensagens_nao_localizado: tuple[str, ...]


PORTAL_TJSP = PortalConfig(
    nome="TJSP - Consulta pública",
    url_consulta=(
        "https://eproc-consulta.tjsp.jus.br/consulta_1g/"
        "externo_controlador.php?"
        "acao=tjsp@consulta_unificada_publica/consultar"
    ),
    campos_processo=(
        "#txtNumProcesso",
        'xpath://input[@name="num_processo"]',
    ),
    botoes_consultar=(
        "#sbmNovo",
        'xpath://button[@type="submit" and '
        'contains(translate(normalize-space(string(.)), '
        '"CONSULTAR", "consultar"), "consultar")]',
        'xpath://input[@type="submit" and '
        'contains(translate(@value, "CONSULTAR", "consultar"), "consultar")]',
    ),
    botoes_voltar=(
        "#btnVoltar",
        'xpath://button[@name="btnVoltar"]',
        'xpath://button[contains(translate(normalize-space(string(.)), '
        '"VOLTAR", "voltar"), "voltar")]',
        'xpath://i[@id="setaVoltar"]/'
        'ancestor::*[self::a or self::button][1]',
        "#setaVoltar",
        'xpath://a[contains(translate(normalize-space(string(.)), '
        '"VOLTAR", "voltar"), "voltar")]',
    ),
    mensagens_captcha=(
        "aguarde a verificação do captcha",
        "aguarde a verificacao do captcha",
        "verificando se você é humano",
        "verificando se voce e humano",
        "checking your browser",
        "código de segurança",
        "codigo de seguranca",
        "não sou um robô",
        "nao sou um robo",
    ),
    mensagens_senha_necessaria=(
        "digite a senha do processo",
        "senha do processo",
        "é necessário informar uma senha",
        "e necessario informar uma senha",
        "chave de acesso do processo",
        "informe a chave de acesso",
        "informe a senha",
    ),
    mensagens_nao_localizado=(
        "processo não localizado na base pública",
        "processo nao localizado na base publica",
        "utilize a pesquisa avançada caso tenha chave/senha de acesso",
        "utilize a pesquisa avancada caso tenha chave/senha de acesso",
    ),
)


# ============================================================
# JAVASCRIPT DE DETECÇÃO/EXTRAÇÃO
# Faz uma única leitura do DOM por ciclo e reconhece:
# 1) eproc: Evento | Data/Hora | Descrição | Usuário
# 2) linhas tr.containerMovimentacao
# 3) tabela genérica: Data + Movimento/Movimentação/Descrição
# ============================================================

JS_ANALISAR_PAGINA = r"""
return (() => {
    const MAX_MOVEMENTS = 3;
    const clean = (value) => String(value || '')
        .replace(/\u00a0/g, ' ')
        .replace(/[ \t]+/g, ' ')
        .replace(/\n\s*\n+/g, '\n')
        .trim();

    const key = (value) => clean(value)
        .normalize('NFD')
        .replace(/[\u0300-\u036f]/g, '')
        .toLowerCase();

    const result = {
        readyState: document.readyState,
        url: location.href,
        title: document.title,
        bodyText: clean(document.body ? document.body.innerText : '').slice(0, 50000),
        layout: null,
        tableFound: false,
        movements: [],
        accessRequired: false,
        captchaFound: false,
        intermediateFound: false,
        intermediateLinks: []
    };

    const bodyKey = key(result.bodyText);
    const visible = (element) => {
        if (!element) return false;
        const style = window.getComputedStyle(element);
        const rect = element.getBoundingClientRect();
        return style
            && style.display !== 'none'
            && style.visibility !== 'hidden'
            && Number(style.opacity || 1) !== 0
            && rect.width > 0
            && rect.height > 0;
    };

    const accessPhrases = [
        'digite a senha do processo',
        'senha do processo',
        'e necessario informar uma senha',
        'é necessário informar uma senha',
        'chave de acesso do processo',
        'informe a chave de acesso',
        'informe a senha'
    ].map(key);

    const accessInputs = Array.from(document.querySelectorAll('input, textarea'))
        .filter((element) => {
            const type = key(element.getAttribute('type'));
            if (type === 'hidden' || element.disabled) return false;
            if (!visible(element)) return false;
            const attrs = key([
                element.getAttribute('type'),
                element.getAttribute('name'),
                element.getAttribute('id'),
                element.getAttribute('aria-label'),
                element.getAttribute('placeholder')
            ].join(' '));
            return type === 'password'
                || attrs.includes('senha')
                || attrs.includes('chave de acesso');
        });

    result.accessRequired = accessInputs.length > 0
        || accessPhrases.some((phrase) => bodyKey.includes(phrase));

    const captchaPhrases = [
        'aguarde a verificacao do captcha',
        'verificando se voce e humano',
        'checking your browser',
        'codigo de seguranca',
        'nao sou um robo',
        'captcha'
    ].map(key);

    const captchaElements = Array.from(document.querySelectorAll(
        'iframe[src*="captcha" i], input[name*="captcha" i], '
        + 'img[src*="captcha" i], [class*="captcha" i], [id*="captcha" i], '
        + '.g-recaptcha, [data-sitekey]'
    )).filter(visible);

    result.captchaFound = captchaElements.length > 0
        || captchaPhrases.some((phrase) => bodyKey.includes(phrase));

    const processNumberRegex = /\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}/;
    result.intermediateLinks = Array.from(document.querySelectorAll('a'))
        .map((anchor, index) => ({
            index,
            text: clean(anchor.innerText),
            href: anchor.href || anchor.getAttribute('href') || '',
            className: anchor.className || ''
        }))
        .filter((link) => {
            const classKey = key(link.className);
            return processNumberRegex.test(link.text)
                || (
                    classKey.includes('linkprocesso')
                    && /show\.do|exibir_processo|processo/i.test(link.href)
                );
        })
        .slice(0, 10);

    const tables = Array.from(document.querySelectorAll('table'));
    const addMovement = (item) => {
        if (item.evento || item.data_hora || item.descricao || item.usuario) {
            result.movements.push(item);
        }
        return result.movements.length >= MAX_MOVEMENTS;
    };

    // Layout eproc.
    const eprocTable = tables.find((table) => {
        const headers = Array.from(table.querySelectorAll('th'))
            .map((th) => key(th.innerText));
        return headers.includes('evento')
            && headers.includes('data/hora')
            && headers.includes('descricao')
            && headers.includes('usuario');
    });

    if (eprocTable) {
        result.layout = 'eproc_eventos';
        result.tableFound = true;

        for (const row of Array.from(eprocTable.rows || [])) {
            const cells = Array.from(row.children || [])
                .filter((cell) => cell.tagName === 'TD');
            if (cells.length < 4) continue;

            const item = {
                evento: clean(cells[0].innerText) || null,
                data_hora: clean(cells[1].innerText) || null,
                descricao: clean(cells[2].innerText),
                usuario: clean(cells[3].innerText) || null
            };

            if (addMovement(item)) break;
        }
        return result;
    }

    // Layout com containerMovimentacao.
    const movementRows = Array.from(
        document.querySelectorAll('.containerMovimentacao')
    ).filter((row) =>
        row.querySelector('.dataMovimentacao')
        || row.querySelector('.descricaoMovimentacao')
    );

    if (movementRows.length) {
        result.layout = 'container_movimentacao';
        result.tableFound = true;

        for (const row of movementRows) {
            const dateCell = row.querySelector('.dataMovimentacao');
            const descriptionCell = row.querySelector('.descricaoMovimentacao');
            const item = {
                evento: null,
                data_hora: clean(dateCell ? dateCell.innerText : '') || null,
                descricao: clean(descriptionCell ? descriptionCell.innerText : ''),
                usuario: null
            };

            if (addMovement(item)) break;
        }
        return result;
    }

    // Fallback genérico por cabeçalhos.
    for (const table of tables) {
        const rows = Array.from(table.rows || []);

        for (let headerIndex = 0; headerIndex < rows.length; headerIndex++) {
            const headerCells = Array.from(rows[headerIndex].cells || []);
            if (!headerCells.some((cell) => cell.tagName === 'TH')) continue;

            const headers = headerCells.map((cell) => key(cell.innerText));
            const dateIndex = headers.findIndex((value) =>
                value === 'data' || value === 'data/hora' || value === 'data hora'
            );
            const descriptionIndex = headers.findIndex((value) =>
                value === 'movimento'
                || value === 'movimentacao'
                || value === 'descricao'
                || value.includes('movimento')
                || value.includes('descricao')
            );
            const eventIndex = headers.findIndex((value) => value === 'evento');
            const userIndex = headers.findIndex((value) => value === 'usuario');

            if (dateIndex < 0 || descriptionIndex < 0) continue;

            result.layout = 'tabela_generica_movimentacoes';
            result.tableFound = true;

            for (let rowIndex = headerIndex + 1; rowIndex < rows.length; rowIndex++) {
                const cells = Array.from(rows[rowIndex].cells || []);
                if (cells.length <= Math.max(dateIndex, descriptionIndex)) continue;

                const item = {
                    evento: eventIndex >= 0 && cells[eventIndex]
                        ? clean(cells[eventIndex].innerText) || null
                        : null,
                    data_hora: clean(cells[dateIndex].innerText) || null,
                    descricao: clean(cells[descriptionIndex].innerText),
                    usuario: userIndex >= 0 && cells[userIndex]
                        ? clean(cells[userIndex].innerText) || null
                        : null
                };

                if (addMovement(item)) break;
            }
            return result;
        }
    }

    result.intermediateFound = result.intermediateLinks.length > 0
        && !result.tableFound;
    return result;
})();
"""


# ============================================================
# LOGS E UTILITÁRIOS
# ============================================================


def configurar_logs() -> logging.Logger:
    PASTA_LOGS.mkdir(parents=True, exist_ok=True)
    PASTA_DEBUG.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger("consulta_processos_tjsp")
    logger.setLevel(logging.DEBUG)
    if logger.handlers:
        return logger

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        "%Y-%m-%d %H:%M:%S",
    )

    terminal = logging.StreamHandler()
    terminal.setLevel(logging.INFO)
    terminal.setFormatter(formatter)

    arquivo = logging.FileHandler(ARQUIVO_LOG, encoding="utf-8")
    arquivo.setLevel(logging.DEBUG)
    arquivo.setFormatter(formatter)

    logger.addHandler(terminal)
    logger.addHandler(arquivo)
    return logger


logger = configurar_logs()


def texto_limpo(value: Any) -> str:
    return " ".join(("" if value is None else str(value)).replace("\xa0", " ").split())


def chave_texto(value: Any) -> str:
    value = unicodedata.normalize("NFD", texto_limpo(value))
    return "".join(c for c in value if unicodedata.category(c) != "Mn").casefold()


def somente_digitos(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", "" if value is None else str(value))


def formatar_cnj(value: Any) -> str:
    digits = somente_digitos(value)
    if len(digits) != 20:
        return texto_limpo(value)
    return (
        f"{digits[:7]}-{digits[7:9]}.{digits[9:13]}."
        f"{digits[13]}.{digits[14:16]}.{digits[16:]}"
    )


def agora_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def pausa(intervalo: tuple[float, float]) -> None:
    time.sleep(random.uniform(*intervalo))


def contem_mensagem(texto: str, mensagens: Sequence[str]) -> str | None:
    page_key = chave_texto(texto)
    for mensagem in mensagens:
        if chave_texto(mensagem) in page_key:
            return mensagem
    return None


def _texto_no_html(elemento: Any) -> str:
    try:
        return texto_limpo(elemento.get_text(" ", strip=True))
    except Exception:
        return texto_limpo(elemento)


def _atributos_chave(elemento: Any, nomes: Sequence[str]) -> str:
    partes: list[str] = []
    for nome in nomes:
        try:
            partes.append(str(elemento.get(nome) or ""))
        except Exception:
            pass
    return chave_texto(" ".join(partes))


def _style_oculta_elemento(style: Any) -> bool:
    style_key = chave_texto(style)
    return (
        "display:none" in style_key.replace(" ", "")
        or "visibility:hidden" in style_key.replace(" ", "")
        or "opacity:0" in style_key.replace(" ", "")
    )


def _elemento_aparentemente_visivel_html(elemento: Any) -> bool:
    current = elemento
    while current is not None and getattr(current, "name", None):
        try:
            if current.has_attr("hidden"):
                return False
            if _style_oculta_elemento(current.get("style")):
                return False
        except Exception:
            pass
        current = getattr(current, "parent", None)
    return True


def _texto_visivel_html(soup: Any) -> str:
    partes: list[str] = []
    for text_node in soup.find_all(string=True):
        parent = getattr(text_node, "parent", None)
        if parent is None or getattr(parent, "name", None) in {"script", "style", "noscript"}:
            continue
        if _elemento_aparentemente_visivel_html(parent):
            partes.append(str(text_node))
    return texto_limpo(" ".join(partes))


def _adicionar_movimentacao(
    movements: list[dict[str, Any]],
    item: dict[str, Any],
) -> bool:
    if item.get("evento") or item.get("data_hora") or item.get("descricao") or item.get("usuario"):
        movements.append(item)
    return len(movements) >= MAX_MOVIMENTACOES


def analisar_html(
    html: str,
    url: str | None = None,
    title: str | None = None,
) -> dict[str, Any]:
    """Analisa HTML estático com a mesma semântica do JS usado no navegador."""
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html or "", "html.parser")
    for node in soup(["script", "style", "noscript"]):
        node.decompose()

    body_text = _texto_visivel_html(soup)
    body_key = chave_texto(body_text)
    result: dict[str, Any] = {
        "readyState": "complete",
        "url": url,
        "title": title or (soup.title.string.strip() if soup.title and soup.title.string else None),
        "bodyText": body_text[:50000],
        "layout": None,
        "tableFound": False,
        "movements": [],
        "accessRequired": False,
        "captchaFound": False,
        "intermediateFound": False,
        "intermediateLinks": [],
    }

    access_phrases = [
        "digite a senha do processo",
        "senha do processo",
        "é necessário informar uma senha",
        "e necessario informar uma senha",
        "chave de acesso do processo",
        "informe a chave de acesso",
        "informe a senha",
    ]
    for field in soup.find_all(["input", "textarea"]):
        type_key = chave_texto(field.get("type"))
        if type_key == "hidden" or field.has_attr("disabled"):
            continue
        if not _elemento_aparentemente_visivel_html(field):
            continue
        attrs = _atributos_chave(
            field,
            ("type", "name", "id", "aria-label", "placeholder"),
        )
        if type_key == "password" or "senha" in attrs or "chave de acesso" in attrs:
            result["accessRequired"] = True
            break
    if not result["accessRequired"]:
        result["accessRequired"] = any(
            chave_texto(phrase) in body_key for phrase in access_phrases
        )

    captcha_phrases = [
        "aguarde a verificação do captcha",
        "aguarde a verificacao do captcha",
        "verificando se você é humano",
        "verificando se voce e humano",
        "checking your browser",
        "código de segurança",
        "codigo de seguranca",
        "não sou um robô",
        "nao sou um robo",
        "captcha",
    ]
    result["captchaFound"] = any(
        chave_texto(phrase) in body_key for phrase in captcha_phrases
    )
    if not result["captchaFound"]:
        for element in soup.find_all(True):
            attrs = _atributos_chave(element, ("id", "class", "name", "src"))
            if "captcha" in attrs or "g-recaptcha" in attrs:
                result["captchaFound"] = True
                break

    process_number_regex = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")
    for index, anchor in enumerate(soup.find_all("a")):
        text = _texto_no_html(anchor)
        href = texto_limpo(anchor.get("href"))
        class_key = chave_texto(" ".join(anchor.get("class", [])))
        if process_number_regex.search(text) or (
            "linkprocesso" in class_key
            and re.search(r"show\.do|exibir_processo|processo", href, re.I)
        ):
            result["intermediateLinks"].append(
                {
                    "index": index,
                    "text": text,
                    "href": href,
                    "className": " ".join(anchor.get("class", [])),
                }
            )
        if len(result["intermediateLinks"]) >= 10:
            break

    tables = soup.find_all("table")
    movements: list[dict[str, Any]] = []

    for table in tables:
        headers = [chave_texto(_texto_no_html(th)) for th in table.find_all("th")]
        if not {
            "evento",
            "data/hora",
            "descricao",
            "usuario",
        }.issubset(set(headers)):
            continue

        result["layout"] = "eproc_eventos"
        result["tableFound"] = True
        for row in table.find_all("tr"):
            cells = row.find_all("td", recursive=False) or row.find_all("td")
            if len(cells) < 4:
                continue
            item = {
                "evento": _texto_no_html(cells[0]) or None,
                "data_hora": _texto_no_html(cells[1]) or None,
                "descricao": _texto_no_html(cells[2]),
                "usuario": _texto_no_html(cells[3]) or None,
            }
            if _adicionar_movimentacao(movements, item):
                break
        result["movements"] = movements
        return result

    movement_rows = soup.select(".containerMovimentacao")
    if movement_rows:
        result["layout"] = "container_movimentacao"
        result["tableFound"] = True
        for row in movement_rows:
            date_cell = row.select_one(".dataMovimentacao")
            description_cell = row.select_one(".descricaoMovimentacao")
            item = {
                "evento": None,
                "data_hora": _texto_no_html(date_cell) or None,
                "descricao": _texto_no_html(description_cell),
                "usuario": None,
            }
            if _adicionar_movimentacao(movements, item):
                break
        result["movements"] = movements
        return result

    for table in tables:
        rows = table.find_all("tr")
        for header_index, row in enumerate(rows):
            header_cells = row.find_all(["th", "td"], recursive=False) or row.find_all(["th", "td"])
            if not any(cell.name == "th" for cell in header_cells):
                continue
            headers = [chave_texto(_texto_no_html(cell)) for cell in header_cells]
            date_index = next(
                (
                    index
                    for index, value in enumerate(headers)
                    if value in {"data", "data/hora", "data hora"}
                ),
                -1,
            )
            description_index = next(
                (
                    index
                    for index, value in enumerate(headers)
                    if value in {"movimento", "movimentacao", "descricao"}
                    or "movimento" in value
                    or "descricao" in value
                ),
                -1,
            )
            if date_index < 0 or description_index < 0:
                continue

            event_index = headers.index("evento") if "evento" in headers else -1
            user_index = headers.index("usuario") if "usuario" in headers else -1
            result["layout"] = "tabela_generica_movimentacoes"
            result["tableFound"] = True

            for data_row in rows[header_index + 1 :]:
                cells = data_row.find_all("td", recursive=False) or data_row.find_all("td")
                if len(cells) <= max(date_index, description_index):
                    continue
                item = {
                    "evento": _texto_no_html(cells[event_index]) if event_index >= 0 and event_index < len(cells) else None,
                    "data_hora": _texto_no_html(cells[date_index]) or None,
                    "descricao": _texto_no_html(cells[description_index]),
                    "usuario": _texto_no_html(cells[user_index]) if user_index >= 0 and user_index < len(cells) else None,
                }
                if _adicionar_movimentacao(movements, item):
                    break
            result["movements"] = movements
            return result

    result["intermediateFound"] = bool(result["intermediateLinks"]) and not result["tableFound"]
    return result


def classificar_pagina(snapshot: dict[str, Any], portal: PortalConfig) -> str:
    page_text = str(snapshot.get("bodyText") or "")
    if snapshot.get("movements"):
        return "resultado"
    if snapshot.get("accessRequired") or contem_mensagem(
        page_text, portal.mensagens_senha_necessaria
    ):
        return "senha_necessaria"
    if snapshot.get("captchaFound") or contem_mensagem(
        page_text, portal.mensagens_captcha
    ):
        return "captcha"
    if contem_mensagem(page_text, portal.mensagens_nao_localizado):
        return "nao_localizado"
    if snapshot.get("tableFound"):
        return "resultado"
    if snapshot.get("intermediateFound") or snapshot.get("intermediateLinks"):
        return "pagina_intermediaria"
    return "desconhecida"


def primeiro_elemento(
    tab: Any,
    seletores: Sequence[str],
    timeout: float = 0.2,
) -> tuple[Any | None, str | None]:
    for seletor in seletores:
        try:
            elemento = tab.ele(seletor, timeout=timeout)
            if elemento:
                return elemento, seletor
        except Exception:
            pass
    return None, None


def analisar_pagina(tab: Any) -> dict[str, Any]:
    try:
        result = tab.run_js(JS_ANALISAR_PAGINA)
        if isinstance(result, dict):
            return result
    except Exception as exc:
        logger.debug("DOM ainda indisponível: %s", exc)

    return {
        "readyState": "indisponível",
        "url": getattr(tab, "url", None),
        "title": getattr(tab, "title", None),
        "bodyText": "",
        "layout": None,
        "tableFound": False,
        "movements": [],
        "accessRequired": False,
        "captchaFound": False,
        "intermediateFound": False,
        "intermediateLinks": [],
    }


class TentativaEncerrada(Exception):
    def __init__(self, resultado: dict[str, Any]):
        super().__init__(resultado.get("mensagem") or resultado.get("status"))
        self.resultado = resultado


def resultado_terminal(
    status: str,
    mensagem: str | None,
    snapshot: dict[str, Any] | None,
    inicio: float,
    pagina_intermediaria: bool = False,
) -> dict[str, Any]:
    snapshot = snapshot or {}
    movements = list(snapshot.get("movements") or [])[:MAX_MOVIMENTACOES]
    if status != "sucesso":
        movements = []
    return {
        "status": status,
        "mensagem": mensagem,
        "layout_movimentacoes": snapshot.get("layout"),
        "quantidade_movimentacoes": len(movements),
        "movimentacoes": movements,
        "url_resultado": snapshot.get("url"),
        "tipo_pagina": classificar_pagina(snapshot, PORTAL_TJSP),
        "pagina_intermediaria_detectada": pagina_intermediaria,
        "tempo_resultado_segundos": round(time.perf_counter() - inicio, 4),
    }


# ============================================================
# EXCEL
# ============================================================


def localizar_coluna(headers: Sequence[Any], column_name: str) -> int:
    target = chave_texto(column_name)
    normalized = [chave_texto(value) for value in headers]

    for index, value in enumerate(normalized):
        if value == target:
            return index
    for index, value in enumerate(normalized):
        if target and target in value:
            return index

    raise ValueError(
        f'Coluna "{column_name}" não encontrada. Cabeçalhos: {list(headers)}'
    )


def carregar_processos() -> list[dict[str, Any]]:
    if not ARQUIVO_EXCEL.exists():
        raise FileNotFoundError(f"Excel não encontrado: {ARQUIVO_EXCEL.resolve()}")

    workbook = load_workbook(ARQUIVO_EXCEL, read_only=True, data_only=True)
    try:
        worksheet = (
            workbook[NOME_ABA_EXCEL]
            if NOME_ABA_EXCEL is not None
            else workbook.active
        )
        header = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if header is None:
            raise ValueError("A planilha está vazia.")

        column_index = localizar_coluna(header, NOME_COLUNA_PROCESSO)
        items: list[dict[str, Any]] = []
        seen: set[str] = set()

        for excel_row, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if column_index >= len(row) or row[column_index] is None:
                continue

            original = texto_limpo(row[column_index])
            process_number = formatar_cnj(row[column_index])
            digits = somente_digitos(process_number)
            key = digits or chave_texto(process_number)

            if not original or key in seen:
                if key in seen:
                    logger.warning("Duplicado ignorado: %s", process_number)
                continue
            seen.add(key)

            valid = len(digits) == 20
            items.append(
                {
                    "linha_excel": excel_row,
                    "valor_original_excel": original,
                    "numero_processo": process_number,
                    "chave": key,
                    "valido": valid,
                    "erro_validacao": None if valid else (
                        "O número não possui 20 dígitos. "
                        "Confirme se a célula do Excel está armazenada como texto."
                    ),
                }
            )

        logger.info(
            "Excel carregado | aba=%s | processos únicos=%d",
            worksheet.title,
            len(items),
        )
        return items
    finally:
        workbook.close()


# ============================================================
# CAPTCHA, FORMULÁRIO E RESULTADO
# ============================================================


def aguardar_formulario(tab: Any, portal: PortalConfig) -> tuple[Any, Any]:
    start = time.perf_counter()
    last_log = start
    last_url: str | None = None

    while time.perf_counter() - start < TIMEOUT_FORMULARIO:
        snapshot = analisar_pagina(tab)
        now = time.perf_counter()
        url_atual = str(snapshot.get("url") or getattr(tab, "url", "") or "")

        if url_atual and url_atual != last_url:
            logger.info("URL após carregar formulário: %s", url_atual)
            last_url = url_atual

        field, field_selector = primeiro_elemento(tab, portal.campos_processo)
        button, button_selector = primeiro_elemento(tab, portal.botoes_consultar)

        if field and button:
            logger.info(
                "Formulário pronto em %.2fs | campo=%s | botão=%s",
                time.perf_counter() - start,
                field_selector,
                button_selector,
            )
            logger.info("URL após formulário pronto: %s", url_atual)
            return field, button

        if now - last_log >= INTERVALO_LOG:
            logger.info(
                "Aguardando formulário | %.2fs | readyState=%s | url=%s",
                now - start,
                snapshot.get("readyState"),
                snapshot.get("url"),
            )
            last_log = now
        time.sleep(INTERVALO_VERIFICACAO)

    raise TentativaEncerrada(
        resultado_terminal(
            "timeout",
            "O formulário não ficou disponível dentro do limite.",
            snapshot if "snapshot" in locals() else None,
            start,
        )
    )


def _texto_elemento_dom(elemento: Any) -> str:
    try:
        return texto_limpo(elemento.text)
    except Exception:
        return ""


def _href_elemento_dom(elemento: Any) -> str:
    try:
        return texto_limpo(elemento.attr("href"))
    except Exception:
        return ""


def clicar_link_intermediario(
    tab: Any,
    process_number: str,
    snapshot: dict[str, Any],
) -> dict[str, Any] | None:
    formatted = texto_limpo(process_number)
    candidates = [
        f'xpath://a[contains(normalize-space(string(.)), "{formatted}")]',
        'xpath://a[contains(@class, "linkProcesso") and contains(@href, "show.do")]',
        'xpath://a[contains(@href, "exibir_processo") and contains(@href, "num_processo=")]',
    ]

    for selector in candidates:
        try:
            link = tab.ele(selector, timeout=0.4)
        except Exception:
            link = None
        if not link:
            continue

        info = {
            "texto": _texto_elemento_dom(link) or formatted,
            "href": _href_elemento_dom(link),
            "seletor": selector,
        }
        logger.info(
            "Clique em link intermediário | texto=%s | href=%s",
            info["texto"],
            info["href"],
        )
        try:
            link.wait.clickable(timeout=3)
        except Exception:
            pass
        link.click()
        logger.info("URL após clique intermediário: %s", getattr(tab, "url", None))
        return info

    links = snapshot.get("intermediateLinks") or []
    if links:
        logger.warning("Página intermediária detectada, mas o link não foi clicável: %s", links[0])
    return None


def aguardar_resultado(
    tab: Any,
    portal: PortalConfig,
    process_number: str,
) -> dict[str, Any]:
    start = time.perf_counter()
    last_log = start
    last_useful_change = start
    last_url: str | None = None
    last_page_type: str | None = None
    empty_table_since: float | None = None
    captcha_start: float | None = None
    intermediate_clicked = False
    pagina_intermediaria = False

    while time.perf_counter() - start < TIMEOUT_RESULTADO:
        snapshot = analisar_pagina(tab)
        now = time.perf_counter()
        page_text = str(snapshot.get("bodyText") or "")
        url_atual = str(snapshot.get("url") or getattr(tab, "url", "") or "")
        page_type = classificar_pagina(snapshot, portal)

        if url_atual and url_atual != last_url:
            logger.info("URL após ação/monitoramento: %s", url_atual)
            last_url = url_atual
            last_useful_change = now

        if page_type != last_page_type:
            logger.info("Tipo de página detectada: %s | url=%s", page_type, url_atual)
            last_page_type = page_type
            if page_type != "desconhecida":
                last_useful_change = now

        if page_type == "senha_necessaria":
            return resultado_terminal(
                "senha_necessaria",
                "Página solicitou senha ou chave de acesso ao processo.",
                snapshot,
                start,
                pagina_intermediaria=pagina_intermediaria,
            )

        if page_type == "nao_localizado":
            return resultado_terminal(
                "nao_localizado",
                contem_mensagem(page_text, portal.mensagens_nao_localizado),
                snapshot,
                start,
                pagina_intermediaria=pagina_intermediaria,
            )

        if page_type == "captcha":
            if captcha_start is None:
                captcha_start = now
                logger.warning("Captcha detectado após a consulta; aguardando até %.0fs.", WATCHDOG_ACAO)
            if now - captcha_start >= WATCHDOG_ACAO:
                return resultado_terminal(
                    "captcha_timeout",
                    "Captcha não resolvido dentro do watchdog de 10s.",
                    snapshot,
                    start,
                    pagina_intermediaria=pagina_intermediaria,
                )

            time.sleep(INTERVALO_VERIFICACAO)
            continue

        captcha_start = None
        movements = snapshot.get("movements") or []
        table_found = bool(snapshot.get("tableFound"))
        layout = snapshot.get("layout")

        if movements:
            movements = movements[:MAX_MOVIMENTACOES]
            snapshot["movements"] = movements
            elapsed = time.perf_counter() - start
            logger.info(
                "Resultado pronto | layout=%s | movimentações extraídas=%d (limite=%d) | %.2fs",
                layout,
                len(movements),
                MAX_MOVIMENTACOES,
                elapsed,
            )
            return resultado_terminal(
                "sucesso",
                None,
                snapshot,
                start,
                pagina_intermediaria=pagina_intermediaria,
            )

        if table_found:
            empty_table_since = empty_table_since or now
            if now - empty_table_since >= 2.0:
                logger.info(
                    "Tabela localizada sem movimentações | layout=%s | limite=%d",
                    layout,
                    MAX_MOVIMENTACOES,
                )
                return resultado_terminal(
                    "sucesso",
                    "Tabela localizada, mas sem movimentações.",
                    snapshot,
                    start,
                    pagina_intermediaria=pagina_intermediaria,
                )
        else:
            empty_table_since = None

        if page_type == "pagina_intermediaria":
            pagina_intermediaria = True
            if not intermediate_clicked:
                link_info = clicar_link_intermediario(tab, process_number, snapshot)
                if not link_info:
                    return resultado_terminal(
                        "pagina_intermediaria",
                        "Página intermediária encontrada, mas nenhum link de processo foi clicável.",
                        snapshot,
                        start,
                        pagina_intermediaria=True,
                    )
                intermediate_clicked = True
                last_useful_change = now
                time.sleep(INTERVALO_VERIFICACAO)
                continue

        if now - last_useful_change >= WATCHDOG_ACAO:
            status = "captcha_timeout" if page_type == "captcha" else "timeout"
            mensagem = (
                "Captcha não resolvido dentro do watchdog de 10s."
                if status == "captcha_timeout"
                else "Nenhuma mudança útil ocorreu dentro do watchdog de 10s."
            )
            return resultado_terminal(
                status,
                mensagem,
                snapshot,
                start,
                pagina_intermediaria=pagina_intermediaria,
            )

        if now - last_log >= INTERVALO_LOG:
            logger.info(
                "Aguardando resultado | %.2fs | tipo=%s | layout=%s | tabela=%s | url=%s",
                now - start,
                page_type,
                layout,
                table_found,
                snapshot.get("url"),
            )
            last_log = now
        time.sleep(INTERVALO_VERIFICACAO)

    return resultado_terminal(
        "timeout",
        "Nenhum resultado reconhecido apareceu dentro do limite.",
        snapshot if "snapshot" in locals() else None,
        start,
        pagina_intermediaria=pagina_intermediaria,
    )


# ============================================================
# ABAS E DIAGNÓSTICO
# ============================================================


def abrir_aba(
    browser: ChromiumPage,
    portal: PortalConfig,
    process_number: str,
    attempt: int,
) -> Any:
    logger.info(
        "Abrindo nova aba limpa | processo=%s | tentativa=%d",
        process_number,
        attempt,
    )
    tab = browser.new_tab()
    tab.get(portal.url_consulta)
    logger.info(
        "Aba aberta | id=%s | processo=%s | url=%s",
        id(tab),
        process_number,
        getattr(tab, "url", None),
    )
    pausa(PAUSA_APOS_ABRIR_ABA)
    return tab


def fechar_aba(
    tab: Any | None,
    process_number: str | None = None,
    attempt: int | None = None,
) -> None:
    if tab is None:
        return
    try:
        logger.info(
            "Fechando aba da tentativa | id=%s | processo=%s | tentativa=%s | url=%s",
            id(tab),
            process_number,
            attempt,
            getattr(tab, "url", None),
        )
        tab.close()
        logger.info("Aba da tentativa fechada | id=%s", id(tab))
    except Exception as exc:
        logger.warning("Falha ao fechar a aba: %s", exc)


def salvar_diagnostico(
    tab: Any,
    process_number: str,
    attempt: int,
    reason: str,
) -> None:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = f"{somente_digitos(process_number) or 'processo'}_t{attempt}_{stamp}"

    try:
        (PASTA_DEBUG / f"{prefix}.html").write_text(str(tab.html), encoding="utf-8")
    except Exception:
        pass

    try:
        (PASTA_DEBUG / f"{prefix}.json").write_text(
            json.dumps(
                {
                    "processo": process_number,
                    "tentativa": attempt,
                    "motivo": reason,
                    "capturado_em": agora_iso(),
                    "url": getattr(tab, "url", None),
                    "snapshot": analisar_pagina(tab),
                },
                ensure_ascii=False,
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )
    except Exception:
        pass

    try:
        tab.get_screenshot(
            path=str(PASTA_DEBUG), name=f"{prefix}.png", full_page=True
        )
    except Exception:
        pass


# ============================================================
# CONSULTA DE UM PROCESSO
# Cada tentativa usa uma aba nova e sempre fecha essa aba no final.
# ============================================================


def executar_tentativa(
    browser: ChromiumPage,
    portal: PortalConfig,
    item: dict[str, Any],
    attempt: int,
) -> dict[str, Any]:
    tab: Any | None = None
    start = time.perf_counter()

    def finalizar(resultado: dict[str, Any]) -> dict[str, Any]:
        resultado.update(
            {
                "numero_processo": item["numero_processo"],
                "linha_excel": item["linha_excel"],
                "valor_original_excel": item["valor_original_excel"],
                "portal": portal.nome,
                "tentativa_utilizada": attempt,
                "consultado_em": agora_iso(),
                "duracao_total_segundos": round(time.perf_counter() - start, 4),
            }
        )
        logger.info(
            "Encerrando tentativa | processo=%s | status=%s | motivo=%s | movimentações=%d",
            item["numero_processo"],
            resultado.get("status"),
            resultado.get("mensagem") or "concluído",
            resultado.get("quantidade_movimentacoes", 0),
        )
        return resultado

    try:
        tab = abrir_aba(browser, portal, item["numero_processo"], attempt)
        field, button = aguardar_formulario(tab, portal)
        pausa(PAUSA_ANTES_DE_PREENCHER)

        try:
            field.click()
            field.clear()
        except Exception:
            pass
        field.input(item["numero_processo"])
        logger.info("URL após preencher processo: %s", getattr(tab, "url", None))

        pausa(PAUSA_ANTES_DE_CONSULTAR)
        try:
            button.wait.clickable(timeout=5)
        except Exception:
            pass
        button.click()
        logger.info("URL após clicar consultar: %s", getattr(tab, "url", None))

        result = aguardar_resultado(tab, portal, item["numero_processo"])
        return finalizar(result)

    except TentativaEncerrada as exc:
        return finalizar(exc.resultado)
    except Exception as exc:
        if tab is not None:
            salvar_diagnostico(
                tab,
                item["numero_processo"],
                attempt,
                f"{type(exc).__name__}: {exc}",
            )
        raise
    finally:
        fechar_aba(tab, item.get("numero_processo"), attempt)


def consultar_processo(
    browser: ChromiumPage,
    portal: PortalConfig,
    item: dict[str, Any],
) -> dict[str, Any]:
    if not item["valido"]:
        return {
            "numero_processo": item["numero_processo"],
            "linha_excel": item["linha_excel"],
            "valor_original_excel": item["valor_original_excel"],
            "portal": portal.nome,
            "status": "numero_invalido",
            "mensagem": item["erro_validacao"],
            "tentativa_utilizada": 0,
            "consultado_em": agora_iso(),
            "quantidade_movimentacoes": 0,
            "movimentacoes": [],
        }

    errors: list[dict[str, Any]] = []

    for attempt in range(1, MAX_TENTATIVAS_POR_PROCESSO + 1):
        logger.info(
            "Tentativa %d/%d | processo=%s",
            attempt,
            MAX_TENTATIVAS_POR_PROCESSO,
            item["numero_processo"],
        )
        try:
            result = executar_tentativa(browser, portal, item, attempt)
            if (
                result.get("status") in STATUS_TRANSITORIOS_TENTAR_NOVAMENTE
                and attempt < MAX_TENTATIVAS_POR_PROCESSO
            ):
                errors.append(
                    {
                        "tentativa": attempt,
                        "tipo": "StatusTransitorio",
                        "status": result.get("status"),
                        "mensagem": result.get("mensagem"),
                        "url_resultado": result.get("url_resultado"),
                        "ocorrido_em": agora_iso(),
                    }
                )
                logger.warning(
                    "Status transitório na tentativa %d: %s | %s. Nova tentativa em aba limpa.",
                    attempt,
                    result.get("status"),
                    result.get("mensagem"),
                )
                pausa(PAUSA_ENTRE_TENTATIVAS)
                continue
            result["erros_tentativas_anteriores"] = errors
            return result
        except Exception as exc:
            errors.append(
                {
                    "tentativa": attempt,
                    "tipo": type(exc).__name__,
                    "mensagem": str(exc),
                    "ocorrido_em": agora_iso(),
                }
            )
            logger.exception("Falha na tentativa %d: %s", attempt, exc)
            if attempt < MAX_TENTATIVAS_POR_PROCESSO:
                pausa(PAUSA_ENTRE_TENTATIVAS)

    return {
        "numero_processo": item["numero_processo"],
        "linha_excel": item["linha_excel"],
        "valor_original_excel": item["valor_original_excel"],
        "portal": portal.nome,
        "status": "erro",
        "mensagem": "Todas as tentativas falharam.",
        "tentativa_utilizada": MAX_TENTATIVAS_POR_PROCESSO,
        "consultado_em": agora_iso(),
        "quantidade_movimentacoes": 0,
        "movimentacoes": [],
        "erros_tentativas": errors,
    }


# ============================================================
# CHECKPOINT, RETOMADA E RESUMO
# ============================================================

STATUS_RESUMO = (
    "sucesso",
    "nao_localizado",
    "numero_invalido",
    "senha_necessaria",
    "captcha_timeout",
    "pagina_intermediaria",
    "timeout",
    "erro",
)

STATUS_DEFINITIVOS = {"sucesso", "nao_localizado", "numero_invalido"}
STATUS_TRANSITORIOS_TENTAR_NOVAMENTE = {
    "timeout",
    "captcha_timeout",
    "pagina_intermediaria",
}


def normalizar_resultado_movimentacoes(item: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(item)
    status = normalized.get("status", "erro")

    if status != "sucesso":
        normalized["quantidade_movimentacoes"] = 0
        normalized["movimentacoes"] = []
        return normalized

    raw_movements = normalized.get("movimentacoes")
    if isinstance(raw_movements, list):
        movements = raw_movements[:MAX_MOVIMENTACOES]
        normalized["movimentacoes"] = movements
        normalized["quantidade_movimentacoes"] = len(movements)
        return normalized

    try:
        quantity = int(normalized.get("quantidade_movimentacoes", 0) or 0)
    except Exception:
        quantity = 0
    normalized["quantidade_movimentacoes"] = min(max(quantity, 0), MAX_MOVIMENTACOES)
    normalized["movimentacoes"] = []
    return normalized


def carregar_checkpoint() -> dict[str, dict[str, Any]]:
    if not RETOMAR_EXECUCAO_ANTERIOR or not ARQUIVO_JSON.exists():
        return {}
    try:
        data = json.loads(ARQUIVO_JSON.read_text(encoding="utf-8"))
        return {
            somente_digitos(item.get("numero_processo"))
            or chave_texto(item.get("numero_processo")): normalizar_resultado_movimentacoes(item)
            for item in data.get("processos", [])
            if isinstance(item, dict) and item.get("numero_processo")
        }
    except Exception as exc:
        logger.warning("Checkpoint anterior ignorado: %s", exc)
        return {}


def resumo(resultados: Sequence[dict[str, Any]]) -> dict[str, int]:
    output = {
        "total_processados": len(resultados),
        **{status: 0 for status in STATUS_RESUMO},
        "total_movimentacoes": 0,
    }
    for item in resultados:
        item = normalizar_resultado_movimentacoes(item)
        status = item.get("status", "erro")
        output[status if status in output else "erro"] += 1
        output["total_movimentacoes"] += int(
            item.get("quantidade_movimentacoes", 0) or 0
        )
    return output


def salvar_checkpoint(
    resultados: Sequence[dict[str, Any]],
    total_excel: int,
    iniciado_em: str,
) -> None:
    normalized_results = [
        normalizar_resultado_movimentacoes(item) for item in resultados
    ]
    payload = {
        "portal": PORTAL_TJSP.nome,
        "arquivo_excel": str(ARQUIVO_EXCEL.resolve()),
        "coluna_processo": NOME_COLUNA_PROCESSO,
        "iniciado_em": iniciado_em,
        "ultima_atualizacao": agora_iso(),
        "total_processos_excel": total_excel,
        "resumo": resumo(normalized_results),
        "processos": normalized_results,
    }

    ARQUIVO_JSON.parent.mkdir(parents=True, exist_ok=True)
    temporary = ARQUIVO_JSON.with_suffix(ARQUIVO_JSON.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    temporary.replace(ARQUIVO_JSON)


# ============================================================
# LOTE
# ============================================================


def criar_navegador() -> ChromiumPage:
    options = ChromiumOptions()
    options.set_argument("--start-maximized")
    return ChromiumPage(addr_or_opts=options)


def consultar_lote() -> list[dict[str, Any]]:
    processes = carregar_processos()
    if not processes:
        raise ValueError("Nenhum processo encontrado no Excel.")

    existing = carregar_checkpoint()
    by_key: dict[str, dict[str, Any]] = {}

    for item in processes:
        saved = existing.get(item["chave"])
        if saved and saved.get("status") in STATUS_DEFINITIVOS:
            by_key[item["chave"]] = saved

    browser: ChromiumPage | None = None
    started_at = agora_iso()

    try:
        browser = criar_navegador()

        # Aba-base: faz a validação inicial e mantém a sessão/cookies.
        browser.get(PORTAL_TJSP.url_consulta)
        aguardar_formulario(browser, PORTAL_TJSP)
        logger.info("Sessão-base pronta. As consultas usarão abas novas.")

        for index, item in enumerate(processes, start=1):
            if item["chave"] in by_key:
                logger.info(
                    "Já processado | %d/%d | %s | status=%s",
                    index,
                    len(processes),
                    item["numero_processo"],
                    by_key[item["chave"]].get("status"),
                )
                continue

            logger.info("=" * 70)
            logger.info(
                "PROCESSO %d/%d | linha=%d | %s",
                index,
                len(processes),
                item["linha_excel"],
                item["numero_processo"],
            )

            by_key[item["chave"]] = consultar_processo(
                browser, PORTAL_TJSP, item
            )

            ordered = [
                by_key[p["chave"]]
                for p in processes
                if p["chave"] in by_key
            ]
            salvar_checkpoint(ordered, len(processes), started_at)

            partial = resumo(ordered)
            logger.info(
                "Checkpoint | processados=%d | sucessos=%d | "
                "não localizados=%d | inválidos=%d | senha=%d | "
                "captcha_timeout=%d | página intermediária=%d | timeouts=%d | erros=%d",
                partial["total_processados"],
                partial["sucesso"],
                partial["nao_localizado"],
                partial["numero_invalido"],
                partial["senha_necessaria"],
                partial["captcha_timeout"],
                partial["pagina_intermediaria"],
                partial["timeout"],
                partial["erro"],
            )

            if index < len(processes):
                pausa(PAUSA_ENTRE_PROCESSOS)

        final_results = [
            by_key[item["chave"]]
            for item in processes
            if item["chave"] in by_key
        ]
        salvar_checkpoint(final_results, len(processes), started_at)
        return final_results

    finally:
        if browser is not None:
            try:
                browser.quit()
                logger.info("Navegador encerrado.")
            except Exception as exc:
                logger.warning("Erro ao encerrar navegador: %s", exc)


if __name__ == "__main__":
    try:
        results = consultar_lote()
        final_summary = resumo(results)

        print("\nResumo final")
        print(f"Processados: {final_summary['total_processados']}")
        print(f"Sucessos: {final_summary['sucesso']}")
        print(f"Não localizados: {final_summary['nao_localizado']}")
        print(f"Números inválidos: {final_summary['numero_invalido']}")
        print(f"Senha necessária: {final_summary['senha_necessaria']}")
        print(f"Captcha timeout: {final_summary['captcha_timeout']}")
        print(f"Página intermediária: {final_summary['pagina_intermediaria']}")
        print(f"Timeouts: {final_summary['timeout']}")
        print(f"Erros: {final_summary['erro']}")
        print(f"Movimentações: {final_summary['total_movimentacoes']}")
        print(f"JSON: {ARQUIVO_JSON.resolve()}")
        print(f"Log: {ARQUIVO_LOG.resolve()}")

    except Exception as exc:
        logger.exception("Falha fatal: %s", exc)
        print(f"\nFalha na execução. Veja o log: {ARQUIVO_LOG.resolve()}")
