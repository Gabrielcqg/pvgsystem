from __future__ import annotations

import json

import pytest

from app.loader import generate_synthetic_fixtures, load_workbooks


@pytest.fixture()
def synthetic_planilhas(tmp_path):
    return generate_synthetic_fixtures(tmp_path)


def _counts(conn) -> dict[str, int]:
    tables = ["parceiros", "contratos", "parcelas", "lancamentos", "custos_fixos", "import_log"]
    with conn.cursor() as cur:
        return {table: cur.execute(f"SELECT count(*) FROM {table}").fetchone()[0] for table in tables}


@pytest.mark.loader
def test_load_05_validate_only_reports_bindings_and_writes_nothing(clean_db, synthetic_planilhas) -> None:
    before = _counts(clean_db)
    report = load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, validate_only=True)
    after = _counts(clean_db)

    assert report.ok
    assert before == after
    assert "JAN.cash" in report.bindings
    assert "CONTRATOS.contracts" in report.bindings
    assert report.cash_deltas["JAN"] == "-7299.00"
    assert any("Custo fixo possivelmente duplicado" in issue.message for issue in report.warnings)


@pytest.mark.loader
def test_load_01_idempotent_commit_duplicates_nothing(clean_db, synthetic_planilhas) -> None:
    first = load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, conn=clean_db, validate_only=False)
    counts_after_first = _counts(clean_db)
    second = load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, conn=clean_db, validate_only=False)
    counts_after_second = _counts(clean_db)

    assert first.ok and second.ok
    assert counts_after_second["contratos"] == counts_after_first["contratos"]
    assert counts_after_second["parcelas"] == counts_after_first["parcelas"]
    assert counts_after_second["lancamentos"] == counts_after_first["lancamentos"]
    assert counts_after_second["custos_fixos"] == counts_after_first["custos_fixos"]
    assert second.counts["linhas_ignoradas_idempotencia"] >= 20


@pytest.mark.loader
def test_load_02_vocabulary_normalization_and_log(clean_db, synthetic_planilhas) -> None:
    report = load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, conn=clean_db, validate_only=False)
    expected = json.loads(synthetic_planilhas.esperado.read_text(encoding="utf-8"))

    assert report.ok
    with clean_db.cursor() as cur:
        cur.execute("SELECT tipo_honorario, count(*) FROM contratos GROUP BY tipo_honorario")
        distribution = dict(cur.fetchall())
        cur.execute("SELECT count(*) FROM import_log WHERE acao = 'mapear_vocabulario'")
        mapping_logs = cur.fetchone()[0]

    for key, value in expected["tipo_honorario_distribution"].items():
        assert distribution[key] == value
    assert mapping_logs >= 15


@pytest.mark.loader
def test_load_03_orphan_instalments_reconciled_and_flagged(clean_db, synthetic_planilhas) -> None:
    load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, conn=clean_db, validate_only=False)

    with clean_db.cursor() as cur:
        cur.execute("SELECT count(*) FROM contratos WHERE revisar AND cliente LIKE 'Órfão Mensal %'")
        stubs = cur.fetchone()[0]
        cur.execute(
            """
            SELECT count(*)
            FROM parcelas p
            JOIN contratos c ON c.id = p.contrato_id
            WHERE c.cliente LIKE 'Órfão Mensal %'
            """
        )
        orphan_parcels = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM import_log WHERE acao = 'auto_criar_contrato_stub'")
        log_rows = cur.fetchone()[0]

    assert stubs == 4
    assert orphan_parcels == 4
    assert log_rows == 4


@pytest.mark.loader
def test_load_04_cash_chain_recomputed_from_entries(clean_db, synthetic_planilhas) -> None:
    load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, conn=clean_db, validate_only=False)
    expected = json.loads(synthetic_planilhas.esperado.read_text(encoding="utf-8"))

    with clean_db.cursor() as cur:
        cur.execute("SELECT mes, saldo_acumulado::text FROM ind_fluxo_mensal WHERE ano = 2026 AND mes IN (1, 2, 3, 12) ORDER BY mes")
        rows = dict(cur.fetchall())

    assert rows[1] == expected["cash_chain"]["JAN"]
    assert rows[2] == expected["cash_chain"]["FEV"]
    assert rows[3] == expected["cash_chain"]["MAR"]
    assert rows[12] == expected["cash_chain"]["DEZ"]


@pytest.mark.loader
def test_load_06_header_alias_resolution_and_missing_required_header(tmp_path, synthetic_planilhas) -> None:
    from openpyxl import load_workbook

    report = load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, validate_only=True)
    assert report.ok
    assert report.bindings["JAN.cash"]["descricao"] == "Histórico"

    broken = tmp_path / "broken_fluxo.xlsx"
    wb = load_workbook(synthetic_planilhas.fluxo)
    ws = wb["JAN"]
    ws.cell(1, 1).value = "Quando"
    wb.save(broken)
    broken_report = load_workbooks(fluxo_path=broken, contratos_path=synthetic_planilhas.contratos, validate_only=True)
    assert not broken_report.ok
    assert "Cabecalhos disponiveis" in broken_report.errors[0].message


@pytest.mark.loader
@pytest.mark.security
def test_load_07_import_log_written_and_insert_only(clean_db, synthetic_planilhas) -> None:
    load_workbooks(fluxo_path=synthetic_planilhas.fluxo, contratos_path=synthetic_planilhas.contratos, conn=clean_db, validate_only=False)

    with clean_db.cursor() as cur:
        cur.execute("SELECT count(*) FROM import_log WHERE severidade IN ('info', 'aviso')")
        assert cur.fetchone()[0] > 25
        cur.execute("SET ROLE authenticated")
        with pytest.raises(Exception, match="permission denied"):
            cur.execute("UPDATE import_log SET acao = 'x'")
        with pytest.raises(Exception, match="permission denied"):
            cur.execute("DELETE FROM import_log")
        cur.execute("RESET ROLE")
